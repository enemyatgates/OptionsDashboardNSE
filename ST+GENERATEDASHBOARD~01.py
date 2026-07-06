"""
generate_comprehensive_dashboard.py
─────────────────────────────────────
NSE FAO Comprehensive Options Dashboard Generator

Integrates:
  • Participant-Wise Open Interest (OI)
  • Participant-Wise Trading Volume (TV)
  • Strike-Wise Options Chain (OC)

Usage:
    python generate_comprehensive_dashboard.py --folder /path/to/data

    The script auto-discovers files in the folder by naming convention:
      • Open Interest : FAOOIYYYYMMDD*.csv  — latest 5 picked
      • Trading Volume: FAOTVYYYYMMDD*.csv  — latest 5 picked
      • Option Chain  : FAOOCYYYYMMDD*.csv  — latest 1 picked

    Output is written to the same folder as FAOCLAUDEYYYYMMDD.xlsx
    where YYYYMMDD is the date of the latest OI file found.

    • TV and OC files are optional; the script runs cleanly without them.
    • 1 to 5 OI files are required (however many exist up to 5).

Output Sheets (up to 14 depending on inputs):
    01. Raw OI Data           — Parsed participant OI, all dates
    02. Raw TV Data           — Parsed participant TV, all dates
    03. Raw Chain Data        — Parsed option chain, all dates (all strikes stacked)
    04. OI Dashboard          — OI metrics × participant × date
    05. TV Dashboard          — TV metrics × participant × date
    06. Net OI & Bias         — Net OI, L/S ratios, % market share
    07. PCR Analysis          — Participant + aggregate OI/volume PCR
    08. OI-TV Efficiency      — OI/TV ratio, net conviction cross-check
    09. Day-on-Day Changes    — Session-to-session delta on all key metrics
    10. Option Chain View     — Strike-wise OI, ΔOI, Volume, IV, LTP × date
    11. Max Pain              — Max pain strike + per-strike pain values × date
    12. OI Walls & Support    — Top call OI (resistance) + put OI (support) × date
    13. IV Skew               — Implied volatility across strikes × date
    14. Sentiment Summary     — Composite signal: participant + chain signals × date
"""

import argparse
import csv
import json
import re
import sys
from datetime import datetime, timezone
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ══════════════════════════════════════════════════════════════════════════════
# PALETTE & STYLE HELPERS
# ══════════════════════════════════════════════════════════════════════════════
C = {
    "hdr_bg":    "1F2937", "hdr_fg":    "FFFFFF",
    "sect_bg":   "374151", "sect_fg":   "F9FAFB",
    "sub_bg":    "4B5563", "sub_fg":    "F3F4F6",
    "label_bg":  "E5E7EB", "label_fg":  "1F2937",
    "alt":       "F3F4F6", "white":     "FFFFFF",
    "green_bg":  "DCFCE7", "green_fg":  "166534",
    "red_bg":    "FEE2E2", "red_fg":    "991B1B",
    "amber_bg":  "FEF3C7", "amber_fg":  "92400E",
    "blue_bg":   "DBEAFE", "blue_fg":   "1E40AF",
    "purple_bg": "EDE9FE", "purple_fg": "5B21B6",
    "call_bg":   "EFF6FF", "put_bg":    "FFF7ED",
    "wall_bg":   "FEF9C3",
    "sep":       "D1D5DB",
}
PARTICIPANTS = ["Client", "DII", "FII", "Pro"]
THIN = Side(style="thin", color=C["sep"])

def _fill(h): return PatternFill("solid", start_color=h, end_color=h)
def _font(bold=False, color="000000", sz=10, italic=False):
    return Font(name="Arial", bold=bold, color=color, size=sz, italic=italic)
def _bdr(): return Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
def _align(h="right", wrap=False):
    return Alignment(horizontal=h, vertical="center", wrap_text=wrap)

def sc(ws, row, col, value=None, bg=None, fg="000000", bold=False,
       align="right", fmt=None, sz=10, wrap=False, italic=False):
    c = ws.cell(row=row, column=col)
    try:
        if value is not None:
            c.value = value
        c.font  = _font(bold=bold, color=fg, sz=sz, italic=italic)
        if bg: c.fill = _fill(bg)
        c.alignment = _align(align, wrap)
        c.border = _bdr()
        if fmt: c.number_format = fmt
    except AttributeError:
        pass
    return c

def title_row(ws, row, text, n_cols, sz=11):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=n_cols)
    sc(ws, row, 1, text, bg=C["hdr_bg"], fg=C["hdr_fg"], bold=True,
       align="center", sz=sz)
    ws.row_dimensions[row].height = 22

def section_row(ws, row, text, n_cols, bg=None, fg=None):
    bg = bg or C["sect_bg"]; fg = fg or C["sect_fg"]
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=n_cols)
    sc(ws, row, 1, text, bg=bg, fg=fg, bold=True, align="left", sz=10)
    ws.row_dimensions[row].height = 16

def sub_header_row(ws, row, labels, col_start, bg=None, fg=None):
    bg = bg or C["sub_bg"]; fg = fg or C["hdr_fg"]
    for i, lbl in enumerate(labels):
        sc(ws, row, col_start + i, lbl, bg=bg, fg=fg, bold=True,
           align="center", wrap=True)
    ws.row_dimensions[row].height = 28

# ══════════════════════════════════════════════════════════════════════════════
# PARTICIPANT CSV PARSING  (OI & TV — identical format)
# ══════════════════════════════════════════════════════════════════════════════
PART_COLS = [
    "Client Type",
    "Future Index Long",  "Future Index Short",
    "Future Stock Long",  "Future Stock Short",
    "Option Index Call Long",  "Option Index Put Long",
    "Option Index Call Short", "Option Index Put Short",
    "Option Stock Call Long",  "Option Stock Put Long",
    "Option Stock Call Short", "Option Stock Put Short",
    "Total Long Contracts",    "Total Short Contracts",
]
DATA_HEADERS = [
    "Date", "Participant",
    "Fut Idx Long", "Fut Idx Short",
    "Fut Stk Long", "Fut Stk Short",
    "Opt Idx Call Long", "Opt Idx Put Long",
    "Opt Idx Call Short","Opt Idx Put Short",
    "Opt Stk Call Long", "Opt Stk Put Long",
    "Opt Stk Call Short","Opt Stk Put Short",
    "Total Long", "Total Short",
]
COL = {  # 1-based column index in raw data sheets
    "FutIdxL": 3,  "FutIdxS": 4,
    "FutStkL": 5,  "FutStkS": 6,
    "OptIdxCL": 7, "OptIdxPL": 8,
    "OptIdxCS": 9, "OptIdxPS": 10,
    "OptStkCL":11, "OptStkPL":12,
    "OptStkCS":13, "OptStkPS":14,
    "TotL":15,     "TotS":16,
}

def _date_from_header(path: str) -> str:
    with open(path, encoding="utf-8-sig", errors="ignore") as f:
        first = f.readline()
    m = re.search(r"as on\s+(\w+\s+\d+)\s*[,\"]*\s*(\d{4})", first, re.IGNORECASE)
    if m: return f"{m.group(1)} {m.group(2)}"
    m2 = re.search(r"(\d{8})", Path(path).stem)
    if m2:
        d = m2.group(1)
        return f"{d[6:8]}-{d[4:6]}-{d[:4]}"
    return Path(path).stem

def parse_participant_csv(path: str) -> tuple[str, pd.DataFrame]:
    date = _date_from_header(path)
    raw = pd.read_csv(path, header=None, skiprows=2)
    if len(raw.columns) == 16:
        raw = raw.iloc[:, :15]
    raw.columns = PART_COLS
    raw = raw[raw["Client Type"].isin(PARTICIPANTS)].copy()
    for col in PART_COLS[1:]:
        raw[col] = pd.to_numeric(raw[col], errors="coerce").fillna(0).astype(int)
    raw.insert(0, "Date", date)
    return date, raw

def load_participant_files(paths: list[str],
                           date_overrides: list[str] = None):
    dates, frames = [], []
    for i, p in enumerate(paths):
        date, df = parse_participant_csv(p)
        if date_overrides and i < len(date_overrides):
            df["Date"] = date_overrides[i]
            date = date_overrides[i]
        dates.append(date)
        frames.append(df)
    return dates, pd.concat(frames, ignore_index=True)

# ══════════════════════════════════════════════════════════════════════════════
# OPTION CHAIN CSV PARSING
# ══════════════════════════════════════════════════════════════════════════════
CHAIN_CALL_COLS = ["C_OI","C_CHNG_OI","C_VOLUME","C_IV","C_LTP","C_CHNG",
                   "C_BID_QTY","C_BID","C_ASK","C_ASK_QTY"]
CHAIN_PUT_COLS  = ["P_BID_QTY","P_BID","P_ASK","P_ASK_QTY","P_CHNG","P_LTP",
                   "P_IV","P_VOLUME","P_CHNG_OI","P_OI"]

def _clean_num(v: str):
    try: return float(v.strip().strip('"').replace(",",""))
    except: return None

def _date_from_chain_header(path: str) -> str:
    with open(path, encoding="utf-8-sig", errors="ignore") as f:
        first = f.readline()
    # option-chain-ED-NIFTY-07-Jul-2026.csv
    m = re.search(r"(\d{2}-\w{3}-\d{4})", Path(path).stem)
    if m:
        return m.group(1)
    m2 = re.search(r"(\d{8})", Path(path).stem)
    if m2:
        d = m2.group(1)
        return f"{d[6:8]}-{d[4:6]}-{d[:4]}"
    return Path(path).stem

def parse_chain_csv(path: str, date_override: str = None) -> tuple[str, pd.DataFrame]:
    date = date_override or _date_from_chain_header(path)
    with open(path, encoding="utf-8-sig", errors="ignore") as f:
        lines = f.readlines()
    rows = []
    for line in lines[2:]:
        fields = list(csv.reader([line.rstrip()]))[0]
        if len(fields) < 12: continue
        strike = _clean_num(fields[11])
        if strike is None: continue
        row = {"Date": date, "Strike": strike}
        for i, col in enumerate(CHAIN_CALL_COLS):
            row[col] = _clean_num(fields[i + 1])
        for i, col in enumerate(CHAIN_PUT_COLS):
            row[col] = _clean_num(fields[i + 12])
        rows.append(row)
    df = pd.DataFrame(rows).sort_values("Strike").reset_index(drop=True)
    df["C_OI"]      = df["C_OI"].fillna(0)
    df["P_OI"]      = df["P_OI"].fillna(0)
    df["C_CHNG_OI"] = df["C_CHNG_OI"].fillna(0)
    df["P_CHNG_OI"] = df["P_CHNG_OI"].fillna(0)
    df["C_VOLUME"]  = df["C_VOLUME"].fillna(0)
    df["P_VOLUME"]  = df["P_VOLUME"].fillna(0)
    return date, df

def load_chain_files(paths: list[str], date_overrides: list[str] = None):
    dates, frames = [], []
    for i, p in enumerate(paths):
        ov = date_overrides[i] if (date_overrides and i < len(date_overrides)) else None
        date, df = parse_chain_csv(p, date_override=ov)
        dates.append(date)
        frames.append(df)
    return dates, pd.concat(frames, ignore_index=True)

# ══════════════════════════════════════════════════════════════════════════════
# DERIVED ANALYTICS (computed in Python, written as values)
# ══════════════════════════════════════════════════════════════════════════════
def compute_max_pain(df: pd.DataFrame) -> float:
    """Return the strike with minimum aggregate OI pain."""
    strikes = df["Strike"].values
    pain = {}
    for s in strikes:
        cp = sum(max(0.0, s - k) * df.loc[df.Strike == k, "C_OI"].values[0] for k in strikes)
        pp = sum(max(0.0, k - s) * df.loc[df.Strike == k, "P_OI"].values[0] for k in strikes)
        pain[s] = cp + pp
    return min(pain, key=pain.get), pain

def compute_pcr(df: pd.DataFrame) -> dict:
    tot_c = df["C_OI"].sum()
    tot_p = df["P_OI"].sum()
    tot_cv = df["C_VOLUME"].sum()
    tot_pv = df["P_VOLUME"].sum()
    return {
        "oi_pcr":  round(tot_p / tot_c, 4) if tot_c else None,
        "vol_pcr": round(tot_pv / tot_cv, 4) if tot_cv else None,
        "total_call_oi": tot_c, "total_put_oi": tot_p,
        "total_call_vol": tot_cv, "total_put_vol": tot_pv,
    }

def compute_walls(df: pd.DataFrame, top_n=10) -> dict:
    call_walls = df.nlargest(top_n, "C_OI")[["Strike","C_OI","C_CHNG_OI","C_VOLUME"]].reset_index(drop=True)
    put_walls  = df.nlargest(top_n, "P_OI")[["Strike","P_OI","P_CHNG_OI","P_VOLUME"]].reset_index(drop=True)
    return {"call_walls": call_walls, "put_walls": put_walls}

def compute_atm(df: pd.DataFrame) -> float:
    """Strike with highest combined OI — proxy for ATM."""
    df2 = df.copy()
    df2["combined"] = df2["C_OI"] + df2["P_OI"]
    return df2.loc[df2["combined"].idxmax(), "Strike"]

# ══════════════════════════════════════════════════════════════════════════════
# RAW DATA SHEETS
# ══════════════════════════════════════════════════════════════════════════════
def write_raw_participant_sheet(wb, sheet_name, dates, df) -> dict:
    ws = wb.create_sheet(sheet_name)
    ws.freeze_panes = "C3"
    n = len(DATA_HEADERS)
    title_row(ws, 1, f"NSE FAO — {sheet_name}", n)
    for ci, h in enumerate(DATA_HEADERS, 1):
        sc(ws, 2, ci, h, bg=C["sub_bg"], fg=C["hdr_fg"],
           bold=True, align="center", wrap=True)
    ws.row_dimensions[2].height = 28

    row_map = {}
    r = 3
    for date in dates:
        for pi, part in enumerate(PARTICIPANTS):
            bg = C["alt"] if pi % 2 == 0 else C["white"]
            row_map[(date, part)] = r
            sub = df[(df["Date"] == date) & (df["Client Type"] == part)]
            vals = [date, part] + ([int(sub.iloc[0][c]) for c in PART_COLS[1:]]
                                   if not sub.empty else [0] * 14)
            for ci, v in enumerate(vals, 1):
                sc(ws, r, ci, v, bg=bg,
                   align="left" if ci <= 2 else "right",
                   fmt="#,##0" if ci > 2 else None)
            ws.row_dimensions[r].height = 14
            r += 1

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 10
    for ci in range(3, n + 1):
        ws.column_dimensions[get_column_letter(ci)].width = 14
    return row_map

def write_raw_chain_sheet(wb, dates, chain_df):
    ws = wb.create_sheet("Raw Chain Data")
    ws.freeze_panes = "C3"
    chain_cols = ["Date","Strike","C_OI","C_CHNG_OI","C_VOLUME","C_IV","C_LTP","C_CHNG",
                  "P_OI","P_CHNG_OI","P_VOLUME","P_IV","P_LTP","P_CHNG"]
    display_hdrs = ["Date","Strike","Call OI","Call ΔOI","Call Vol","Call IV","Call LTP",
                    "Call Chng","Put OI","Put ΔOI","Put Vol","Put IV","Put LTP","Put Chng"]
    n = len(display_hdrs)
    title_row(ws, 1, "NSE FAO — Raw Option Chain Data", n)

    # Call / Put super-header
    ws.merge_cells("A2:B2")
    sc(ws, 2, 1, "", bg=C["hdr_bg"])
    ws.merge_cells("C2:I2")
    sc(ws, 2, 3, "CALLS", bg=C["call_bg"], fg=C["blue_fg"],
       bold=True, align="center")
    for off in range(1, 7):
        ws.cell(row=2, column=3+off).fill = _fill(C["call_bg"])
        ws.cell(row=2, column=3+off).border = _bdr()
    ws.merge_cells("J2:N2")
    sc(ws, 2, 10, "PUTS", bg=C["put_bg"], fg=C["amber_fg"],
       bold=True, align="center")
    for off in range(1, 5):
        ws.cell(row=2, column=10+off).fill = _fill(C["put_bg"])
        ws.cell(row=2, column=10+off).border = _bdr()
    ws.row_dimensions[2].height = 16

    for ci, h in enumerate(display_hdrs, 1):
        bg = C["call_bg"] if 3 <= ci <= 9 else (C["put_bg"] if ci >= 10 else C["sub_bg"])
        fg = C["blue_fg"] if 3 <= ci <= 9 else (C["amber_fg"] if ci >= 10 else C["hdr_fg"])
        sc(ws, 3, ci, h, bg=bg, fg=fg, bold=True, align="center", wrap=True)
    ws.row_dimensions[3].height = 28

    r = 4
    for date in dates:
        sub = chain_df[chain_df["Date"] == date].reset_index(drop=True)
        # Section row for date
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=n)
        sc(ws, r, 1, date, bg=C["sect_bg"], fg=C["sect_fg"],
           bold=True, align="left")
        ws.row_dimensions[r].height = 14
        r += 1
        for _, row_data in sub.iterrows():
            bg = C["alt"] if r % 2 == 0 else C["white"]
            for ci, col in enumerate(chain_cols, 1):
                v = row_data.get(col)
                if v is not None and not pd.isna(v):
                    fmt = ("#,##0" if col in ("C_OI","C_CHNG_OI","C_VOLUME","P_OI","P_CHNG_OI","P_VOLUME")
                           else "0.00" if col in ("C_IV","P_IV","C_LTP","P_LTP","C_CHNG","P_CHNG")
                           else "#,##0" if col == "Strike" else None)
                    sc(ws, r, ci, v, bg=bg, align="right", fmt=fmt)
                else:
                    sc(ws, r, ci, None, bg=bg)
            ws.row_dimensions[r].height = 13
            r += 1

    widths = [12, 10] + [13]*7 + [13]*5
    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

# ══════════════════════════════════════════════════════════════════════════════
# PARTICIPANT DASHBOARD FRAMEWORK
# ══════════════════════════════════════════════════════════════════════════════
def ref(sheet, row_map, date, part, col_key):
    r = row_map[(date, part)]
    idx = COL[col_key] if isinstance(col_key, str) else col_key
    return f"'{sheet}'!{get_column_letter(idx)}{r}"

def write_participant_dashboard(wb, sheet_name, dates, row_map, data_sheet, metric_groups):
    ws = wb.create_sheet(sheet_name)
    ws.freeze_panes = "B4"
    nd = len(dates)
    np_ = len(PARTICIPANTS)
    def dc(di): return 2 + di * (np_ + 1)
    total_cols = 1 + nd * (np_ + 1)

    title_row(ws, 1, f"NSE FAO — {sheet_name}", total_cols)
    sc(ws, 2, 1, "Metric", bg=C["hdr_bg"], fg=C["hdr_fg"], bold=True, align="center")
    for di, date in enumerate(dates):
        cs = dc(di)
        ws.merge_cells(start_row=2, start_column=cs, end_row=2, end_column=cs + np_ - 1)
        sc(ws, 2, cs, date, bg=C["hdr_bg"], fg=C["hdr_fg"], bold=True, align="center")
        for off in range(1, np_):
            ws.cell(row=2, column=cs+off).fill = _fill(C["hdr_bg"])
            ws.cell(row=2, column=cs+off).border = _bdr()
        ws.cell(row=2, column=cs+np_).fill = _fill(C["hdr_bg"])
    sc(ws, 3, 1, "", bg=C["sect_bg"])
    for di in range(nd):
        cs = dc(di)
        for pi, part in enumerate(PARTICIPANTS):
            sc(ws, 3, cs+pi, part, bg=C["sect_bg"], fg=C["sect_fg"],
               bold=True, align="center")
        ws.cell(row=3, column=cs+np_).fill = _fill(C["sect_bg"])
    ws.row_dimensions[2].height = 16; ws.row_dimensions[3].height = 16

    r = 4
    for group_label, metrics in metric_groups:
        section_row(ws, r, group_label, total_cols)
        r += 1
        for mi, (label, fn, fmt) in enumerate(metrics):
            bg = C["alt"] if mi % 2 == 0 else C["white"]
            sc(ws, r, 1, label, bg=C["label_bg"], fg=C["label_fg"], align="left")
            for di, date in enumerate(dates):
                cs = dc(di)
                for pi, part in enumerate(PARTICIPANTS):
                    sc(ws, r, cs+pi, fn(date, part, data_sheet, row_map),
                       bg=bg, align="right", fmt=fmt)
                ws.cell(row=r, column=cs+np_).fill = _fill(bg)
            ws.row_dimensions[r].height = 14
            r += 1

    ws.column_dimensions["A"].width = 26
    for di in range(nd):
        cs = dc(di)
        for pi in range(np_): ws.column_dimensions[get_column_letter(cs+pi)].width = 14
        ws.column_dimensions[get_column_letter(cs+np_)].width = 2

# ══════════════════════════════════════════════════════════════════════════════
# SHEET BUILDERS — PARTICIPANT BASED
# ══════════════════════════════════════════════════════════════════════════════
def build_oi_dashboard(wb, dates, oi_row_map):
    DS = "Raw OI Data"
    def R(d,p,k): return ref(DS, oi_row_map, d, p, k)
    def f(k): return lambda d,p,ds,rm: f"={R(d,p,k)}"
    def net(k1,k2): return lambda d,p,ds,rm: f"={R(d,p,k1)}-{R(d,p,k2)}"
    def ls(k1,k2): return lambda d,p,ds,rm: f"=IFERROR({R(d,p,k1)}/{R(d,p,k2)},0)"
    def pct(k): return lambda d,p,ds,rm: (
        f"=IFERROR({R(d,p,k)}/SUM({','.join(ref(DS,oi_row_map,d,pp,k) for pp in PARTICIPANTS)}),0)")

    groups = [
        ("FUTURES — INDEX", [
            ("Long OI",          f("FutIdxL"),          "#,##0"),
            ("Short OI",         f("FutIdxS"),          "#,##0"),
            ("Net OI (L−S)",     net("FutIdxL","FutIdxS"), "#,##0;[Red]-#,##0;-"),
            ("L/S Ratio",        ls("FutIdxL","FutIdxS"),  "0.00x"),
        ]),
        ("FUTURES — STOCK", [
            ("Long OI",          f("FutStkL"),          "#,##0"),
            ("Short OI",         f("FutStkS"),          "#,##0"),
            ("Net OI (L−S)",     net("FutStkL","FutStkS"), "#,##0;[Red]-#,##0;-"),
            ("L/S Ratio",        ls("FutStkL","FutStkS"),  "0.00x"),
        ]),
        ("OPTIONS — INDEX CALLS", [
            ("Call Long OI",     f("OptIdxCL"),         "#,##0"),
            ("Call Short OI",    f("OptIdxCS"),         "#,##0"),
            ("Net Call OI",      net("OptIdxCL","OptIdxCS"), "#,##0;[Red]-#,##0;-"),
        ]),
        ("OPTIONS — INDEX PUTS", [
            ("Put Long OI",      f("OptIdxPL"),         "#,##0"),
            ("Put Short OI",     f("OptIdxPS"),         "#,##0"),
            ("Net Put OI",       net("OptIdxPL","OptIdxPS"), "#,##0;[Red]-#,##0;-"),
        ]),
        ("OPTIONS — STOCK CALLS", [
            ("Call Long OI",     f("OptStkCL"),         "#,##0"),
            ("Call Short OI",    f("OptStkCS"),         "#,##0"),
            ("Net Call OI",      net("OptStkCL","OptStkCS"), "#,##0;[Red]-#,##0;-"),
        ]),
        ("OPTIONS — STOCK PUTS", [
            ("Put Long OI",      f("OptStkPL"),         "#,##0"),
            ("Put Short OI",     f("OptStkPS"),         "#,##0"),
            ("Net Put OI",       net("OptStkPL","OptStkPS"), "#,##0;[Red]-#,##0;-"),
        ]),
        ("TOTAL POSITION", [
            ("Total Long",       f("TotL"),             "#,##0"),
            ("Total Short",      f("TotS"),             "#,##0"),
            ("Net Total OI",     net("TotL","TotS"),    "#,##0;[Red]-#,##0;-"),
            ("L/S Ratio",        ls("TotL","TotS"),     "0.00x"),
            ("% of Mkt Long",    pct("TotL"),           "0.0%"),
            ("% of Mkt Short",   pct("TotS"),           "0.0%"),
        ]),
    ]
    write_participant_dashboard(wb, "OI Dashboard", dates, oi_row_map, DS, groups)

def build_tv_dashboard(wb, dates, tv_row_map):
    DS = "Raw TV Data"
    def R(d,p,k): return ref(DS, tv_row_map, d, p, k)
    def f(k): return lambda d,p,ds,rm: f"={R(d,p,k)}"
    def net(k1,k2): return lambda d,p,ds,rm: f"={R(d,p,k1)}-{R(d,p,k2)}"
    def ls(k1,k2): return lambda d,p,ds,rm: f"=IFERROR({R(d,p,k1)}/{R(d,p,k2)},0)"
    def pct(k): return lambda d,p,ds,rm: (
        f"=IFERROR({R(d,p,k)}/SUM({','.join(ref(DS,tv_row_map,d,pp,k) for pp in PARTICIPANTS)}),0)")

    groups = [
        ("FUTURES — INDEX", [
            ("Buy Volume",       f("FutIdxL"),          "#,##0"),
            ("Sell Volume",      f("FutIdxS"),          "#,##0"),
            ("Net Volume (B−S)", net("FutIdxL","FutIdxS"), "#,##0;[Red]-#,##0;-"),
        ]),
        ("FUTURES — STOCK", [
            ("Buy Volume",       f("FutStkL"),          "#,##0"),
            ("Sell Volume",      f("FutStkS"),          "#,##0"),
            ("Net Volume (B−S)", net("FutStkL","FutStkS"), "#,##0;[Red]-#,##0;-"),
        ]),
        ("OPTIONS — INDEX", [
            ("Call Buy Volume",  f("OptIdxCL"),         "#,##0"),
            ("Put Buy Volume",   f("OptIdxPL"),         "#,##0"),
            ("Volume PCR",       lambda d,p,ds,rm: f"=IFERROR({R(d,p,'OptIdxPL')}/{R(d,p,'OptIdxCL')},0)", "0.00"),
            ("Call Sell Volume", f("OptIdxCS"),         "#,##0"),
            ("Put Sell Volume",  f("OptIdxPS"),         "#,##0"),
        ]),
        ("OPTIONS — STOCK", [
            ("Call Buy Volume",  f("OptStkCL"),         "#,##0"),
            ("Put Buy Volume",   f("OptStkPL"),         "#,##0"),
            ("Volume PCR",       lambda d,p,ds,rm: f"=IFERROR({R(d,p,'OptStkPL')}/{R(d,p,'OptStkCL')},0)", "0.00"),
            ("Call Sell Volume", f("OptStkCS"),         "#,##0"),
            ("Put Sell Volume",  f("OptStkPS"),         "#,##0"),
        ]),
        ("TOTAL VOLUME", [
            ("Total Buy",        f("TotL"),             "#,##0"),
            ("Total Sell",       f("TotS"),             "#,##0"),
            ("Net Volume",       net("TotL","TotS"),    "#,##0;[Red]-#,##0;-"),
            ("% of Market Buy",  pct("TotL"),           "0.0%"),
        ]),
    ]
    write_participant_dashboard(wb, "TV Dashboard", dates, tv_row_map, DS, groups)

def build_net_oi_sheet(wb, dates, oi_row_map):
    DS = "Raw OI Data"
    def R(d,p,k): return ref(DS, oi_row_map, d, p, k)
    def pct(k): return lambda d,p: (
        f"=IFERROR({R(d,p,k)}/SUM({','.join(ref(DS,oi_row_map,d,pp,k) for pp in PARTICIPANTS)}),0)")
    groups = [
        ("NET OPEN INTEREST  (Long − Short)", [
            ("Net Fut Index OI",   lambda d,p: f"={R(d,p,'FutIdxL')}-{R(d,p,'FutIdxS')}", "#,##0;[Red]-#,##0;-"),
            ("Net Fut Stock OI",   lambda d,p: f"={R(d,p,'FutStkL')}-{R(d,p,'FutStkS')}", "#,##0;[Red]-#,##0;-"),
            ("Net Futures OI",     lambda d,p: f"=({R(d,p,'FutIdxL')}+{R(d,p,'FutStkL')})-({R(d,p,'FutIdxS')}+{R(d,p,'FutStkS')})", "#,##0;[Red]-#,##0;-"),
            ("Net Idx Call OI",    lambda d,p: f"={R(d,p,'OptIdxCL')}-{R(d,p,'OptIdxCS')}", "#,##0;[Red]-#,##0;-"),
            ("Net Idx Put OI",     lambda d,p: f"={R(d,p,'OptIdxPL')}-{R(d,p,'OptIdxPS')}", "#,##0;[Red]-#,##0;-"),
            ("Net Stk Call OI",    lambda d,p: f"={R(d,p,'OptStkCL')}-{R(d,p,'OptStkCS')}", "#,##0;[Red]-#,##0;-"),
            ("Net Stk Put OI",     lambda d,p: f"={R(d,p,'OptStkPL')}-{R(d,p,'OptStkPS')}", "#,##0;[Red]-#,##0;-"),
            ("Net Total OI",       lambda d,p: f"={R(d,p,'TotL')}-{R(d,p,'TotS')}", "#,##0;[Red]-#,##0;-"),
        ]),
        ("LONG-SHORT RATIOS", [
            ("Fut Index L/S",      lambda d,p: f"=IFERROR({R(d,p,'FutIdxL')}/{R(d,p,'FutIdxS')},0)", "0.00x"),
            ("Fut Stock L/S",      lambda d,p: f"=IFERROR({R(d,p,'FutStkL')}/{R(d,p,'FutStkS')},0)", "0.00x"),
            ("Total L/S Ratio",    lambda d,p: f"=IFERROR({R(d,p,'TotL')}/{R(d,p,'TotS')},0)", "0.00x"),
        ]),
        ("MARKET SHARE", [
            ("% Mkt Long",         pct("TotL"),         "0.0%"),
            ("% Mkt Short",        pct("TotS"),         "0.0%"),
            ("% Mkt Fut Idx Long", pct("FutIdxL"),      "0.0%"),
            ("% Mkt Fut Stk Short",pct("FutStkS"),      "0.0%"),
        ]),
    ]
    # custom write (same pattern but fn takes only d,p)
    ws = wb.create_sheet("Net OI & Bias")
    ws.freeze_panes = "B4"
    nd = len(dates); np_ = len(PARTICIPANTS)
    def dc(di): return 2 + di * (np_ + 1)
    total_cols = 1 + nd * (np_ + 1)
    title_row(ws, 1, "NSE FAO — Net OI & Market Bias by Participant", total_cols)
    sc(ws, 2, 1, "Metric", bg=C["hdr_bg"], fg=C["hdr_fg"], bold=True, align="center")
    for di, date in enumerate(dates):
        cs = dc(di)
        ws.merge_cells(start_row=2, start_column=cs, end_row=2, end_column=cs+np_-1)
        sc(ws, 2, cs, date, bg=C["hdr_bg"], fg=C["hdr_fg"], bold=True, align="center")
        for off in range(1, np_+1):
            ws.cell(row=2, column=cs+off).fill = _fill(C["hdr_bg"])
            ws.cell(row=2, column=cs+off).border = _bdr()
    sc(ws, 3, 1, "", bg=C["sect_bg"])
    for di in range(nd):
        cs = dc(di)
        for pi, part in enumerate(PARTICIPANTS):
            sc(ws, 3, cs+pi, part, bg=C["sect_bg"], fg=C["sect_fg"], bold=True, align="center")
        ws.cell(row=3, column=cs+np_).fill = _fill(C["sect_bg"])
    ws.row_dimensions[2].height = 16; ws.row_dimensions[3].height = 16
    r = 4
    for group_label, metrics in groups:
        section_row(ws, r, group_label, total_cols); r += 1
        for mi, (label, fn, fmt) in enumerate(metrics):
            bg = C["alt"] if mi % 2 == 0 else C["white"]
            sc(ws, r, 1, label, bg=C["label_bg"], fg=C["label_fg"], align="left")
            for di, date in enumerate(dates):
                cs = dc(di)
                for pi, part in enumerate(PARTICIPANTS):
                    sc(ws, r, cs+pi, fn(date, part), bg=bg, fmt=fmt, align="right")
                ws.cell(row=r, column=cs+np_).fill = _fill(bg)
            ws.row_dimensions[r].height = 14; r += 1
    ws.column_dimensions["A"].width = 26
    for di in range(nd):
        cs = dc(di)
        for pi in range(np_): ws.column_dimensions[get_column_letter(cs+pi)].width = 14
        ws.column_dimensions[get_column_letter(cs+np_)].width = 2

def build_pcr_sheet(wb, dates, oi_row_map, tv_row_map, chain_analytics):
    ws = wb.create_sheet("PCR Analysis")
    ws.freeze_panes = "B4"
    OI = "Raw OI Data"; TV = "Raw TV Data"
    nd = len(dates); np_ = len(PARTICIPANTS)
    def dc(di): return 2 + di * (np_ + 1)
    total_cols = 1 + nd * (np_ + 1)
    title_row(ws, 1, "NSE FAO — Put-Call Ratio Analysis (Participant OI, Volume & Chain)", total_cols)
    sc(ws, 2, 1, "Metric", bg=C["hdr_bg"], fg=C["hdr_fg"], bold=True, align="center")
    for di, date in enumerate(dates):
        cs = dc(di)
        ws.merge_cells(start_row=2, start_column=cs, end_row=2, end_column=cs+np_-1)
        sc(ws, 2, cs, date, bg=C["hdr_bg"], fg=C["hdr_fg"], bold=True, align="center")
        for off in range(1, np_+1):
            ws.cell(row=2, column=cs+off).fill = _fill(C["hdr_bg"])
            ws.cell(row=2, column=cs+off).border = _bdr()
    sc(ws, 3, 1, "", bg=C["sect_bg"])
    for di in range(nd):
        cs = dc(di)
        for pi, part in enumerate(PARTICIPANTS):
            sc(ws, 3, cs+pi, part, bg=C["sect_bg"], fg=C["sect_fg"], bold=True, align="center")
        ws.cell(row=3, column=cs+np_).fill = _fill(C["sect_bg"])
    ws.row_dimensions[2].height = 16; ws.row_dimensions[3].height = 16

    def R(d,p,k): return ref(OI, oi_row_map, d, p, k)
    pcr_groups = [
        ("PARTICIPANT OI PCR  (Put Long / Call Long)", [
            ("Index OI PCR",       lambda d,p: f"=IFERROR({R(d,p,'OptIdxPL')}/{R(d,p,'OptIdxCL')},0)", "0.00"),
            ("Stock OI PCR",       lambda d,p: f"=IFERROR({R(d,p,'OptStkPL')}/{R(d,p,'OptStkCL')},0)", "0.00"),
            ("Combined OI PCR",    lambda d,p: f"=IFERROR(({R(d,p,'OptIdxPL')}+{R(d,p,'OptStkPL')})/({R(d,p,'OptIdxCL')}+{R(d,p,'OptStkCL')}),0)", "0.00"),
            ("Index Net PCR",      lambda d,p: f"=IFERROR(({R(d,p,'OptIdxPL')}-{R(d,p,'OptIdxPS')})/({R(d,p,'OptIdxCL')}-{R(d,p,'OptIdxCS')}),0)", "0.00"),
            ("Short-side PCR",     lambda d,p: f"=IFERROR({R(d,p,'OptIdxPS')}/{R(d,p,'OptIdxCS')},0)", "0.00"),
        ]),
    ]
    if tv_row_map:
        def TV_(d,p,k): return ref(TV, tv_row_map, d, p, k)
        pcr_groups.append(("PARTICIPANT VOLUME PCR  (Put Buy / Call Buy)", [
            ("Index Vol PCR",      lambda d,p: f"=IFERROR({TV_(d,p,'OptIdxPL')}/{TV_(d,p,'OptIdxCL')},0)", "0.00"),
            ("Stock Vol PCR",      lambda d,p: f"=IFERROR({TV_(d,p,'OptStkPL')}/{TV_(d,p,'OptStkCL')},0)", "0.00"),
            ("Combined Vol PCR",   lambda d,p: f"=IFERROR(({TV_(d,p,'OptIdxPL')}+{TV_(d,p,'OptStkPL')})/({TV_(d,p,'OptIdxCL')}+{TV_(d,p,'OptStkCL')}),0)", "0.00"),
        ]))

    r = 4
    for group_label, metrics in pcr_groups:
        section_row(ws, r, group_label, total_cols); r += 1
        for mi, (label, fn, fmt) in enumerate(metrics):
            bg = C["alt"] if mi % 2 == 0 else C["white"]
            sc(ws, r, 1, label, bg=C["label_bg"], fg=C["label_fg"], align="left")
            for di, date in enumerate(dates):
                cs = dc(di)
                for pi, part in enumerate(PARTICIPANTS):
                    sc(ws, r, cs+pi, fn(date, part), bg=bg, fmt=fmt, align="right")
                ws.cell(row=r, column=cs+np_).fill = _fill(bg)
            ws.row_dimensions[r].height = 14; r += 1

    # Chain-level aggregate PCR (single row per date, all participants merged)
    if chain_analytics:
        section_row(ws, r, "CHAIN-LEVEL AGGREGATE PCR  (all strikes, market-wide)", total_cols)
        r += 1
        chain_metrics = [
            ("Aggregate OI PCR (Put OI / Call OI)", "oi_pcr"),
            ("Aggregate Volume PCR (Put Vol / Call Vol)", "vol_pcr"),
        ]
        for mi, (label, key) in enumerate(chain_metrics):
            bg = C["alt"] if mi % 2 == 0 else C["white"]
            sc(ws, r, 1, label, bg=C["label_bg"], fg=C["label_fg"], align="left")
            for di, date in enumerate(dates):
                cs = dc(di)
                val = chain_analytics.get(date, {}).get(key)
                ws.merge_cells(start_row=r, start_column=cs, end_row=r, end_column=cs+np_-1)
                sc(ws, r, cs, val, bg=C["blue_bg"], fg=C["blue_fg"],
                   bold=True, align="center", fmt="0.000")
                for off in range(1, np_):
                    ws.cell(row=r, column=cs+off).fill = _fill(C["blue_bg"])
                    ws.cell(row=r, column=cs+off).border = _bdr()
                ws.cell(row=r, column=cs+np_).fill = _fill(bg)
            ws.row_dimensions[r].height = 14; r += 1

    ws.column_dimensions["A"].width = 38
    for di in range(nd):
        cs = dc(di)
        for pi in range(np_): ws.column_dimensions[get_column_letter(cs+pi)].width = 13
        ws.column_dimensions[get_column_letter(cs+np_)].width = 2

def build_efficiency_sheet(wb, dates, oi_row_map, tv_row_map):
    if not tv_row_map: return
    DS_OI = "Raw OI Data"; DS_TV = "Raw TV Data"
    def R(d,p,k): return ref(DS_OI, oi_row_map, d, p, k)
    def TV_(d,p,k): return ref(DS_TV, tv_row_map, d, p, k)
    groups = [
        ("OI / VOLUME RATIO  (high = positions held with conviction, low = churn)", [
            ("Fut Index OI/TV",    lambda d,p: f"=IFERROR({R(d,p,'FutIdxL')}/{TV_(d,p,'FutIdxL')},0)", "0.00"),
            ("Fut Stock OI/TV",    lambda d,p: f"=IFERROR({R(d,p,'FutStkL')}/{TV_(d,p,'FutStkL')},0)", "0.00"),
            ("Idx Call OI/TV",     lambda d,p: f"=IFERROR({R(d,p,'OptIdxCL')}/{TV_(d,p,'OptIdxCL')},0)", "0.00"),
            ("Idx Put OI/TV",      lambda d,p: f"=IFERROR({R(d,p,'OptIdxPL')}/{TV_(d,p,'OptIdxPL')},0)", "0.00"),
            ("Stk Call OI/TV",     lambda d,p: f"=IFERROR({R(d,p,'OptStkCL')}/{TV_(d,p,'OptStkCL')},0)", "0.00"),
            ("Stk Put OI/TV",      lambda d,p: f"=IFERROR({R(d,p,'OptStkPL')}/{TV_(d,p,'OptStkPL')},0)", "0.00"),
            ("Total OI/TV",        lambda d,p: f"=IFERROR({R(d,p,'TotL')}/{TV_(d,p,'TotL')},0)", "0.00"),
        ]),
        ("NET CONVICTION  (OI direction vs Volume direction — aligned = strong signal)", [
            ("Net Fut Idx OI",     lambda d,p: f"={R(d,p,'FutIdxL')}-{R(d,p,'FutIdxS')}", "#,##0;[Red]-#,##0;-"),
            ("Net Fut Idx Vol",    lambda d,p: f"={TV_(d,p,'FutIdxL')}-{TV_(d,p,'FutIdxS')}", "#,##0;[Red]-#,##0;-"),
            ("Net Idx Call OI",    lambda d,p: f"={R(d,p,'OptIdxCL')}-{R(d,p,'OptIdxCS')}", "#,##0;[Red]-#,##0;-"),
            ("Net Idx Call Vol",   lambda d,p: f"={TV_(d,p,'OptIdxCL')}-{TV_(d,p,'OptIdxCS')}", "#,##0;[Red]-#,##0;-"),
            ("Net Idx Put OI",     lambda d,p: f"={R(d,p,'OptIdxPL')}-{R(d,p,'OptIdxPS')}", "#,##0;[Red]-#,##0;-"),
            ("Net Idx Put Vol",    lambda d,p: f"={TV_(d,p,'OptIdxPL')}-{TV_(d,p,'OptIdxPS')}", "#,##0;[Red]-#,##0;-"),
        ]),
    ]
    ws = wb.create_sheet("OI-TV Efficiency")
    ws.freeze_panes = "B4"
    nd = len(dates); np_ = len(PARTICIPANTS)
    def dc(di): return 2 + di * (np_ + 1)
    total_cols = 1 + nd * (np_ + 1)
    title_row(ws, 1, "NSE FAO — OI / Trading Volume Efficiency Ratio", total_cols)
    sc(ws, 2, 1, "Metric", bg=C["hdr_bg"], fg=C["hdr_fg"], bold=True, align="center")
    for di, date in enumerate(dates):
        cs = dc(di)
        ws.merge_cells(start_row=2, start_column=cs, end_row=2, end_column=cs+np_-1)
        sc(ws, 2, cs, date, bg=C["hdr_bg"], fg=C["hdr_fg"], bold=True, align="center")
        for off in range(1, np_+1):
            ws.cell(row=2, column=cs+off).fill = _fill(C["hdr_bg"])
            ws.cell(row=2, column=cs+off).border = _bdr()
    sc(ws, 3, 1, "", bg=C["sect_bg"])
    for di in range(nd):
        cs = dc(di)
        for pi, part in enumerate(PARTICIPANTS):
            sc(ws, 3, cs+pi, part, bg=C["sect_bg"], fg=C["sect_fg"], bold=True, align="center")
        ws.cell(row=3, column=cs+np_).fill = _fill(C["sect_bg"])
    ws.row_dimensions[2].height = 16; ws.row_dimensions[3].height = 16
    r = 4
    for group_label, metrics in groups:
        section_row(ws, r, group_label, total_cols); r += 1
        for mi, (label, fn, fmt) in enumerate(metrics):
            bg = C["alt"] if mi % 2 == 0 else C["white"]
            sc(ws, r, 1, label, bg=C["label_bg"], fg=C["label_fg"], align="left")
            for di, date in enumerate(dates):
                cs = dc(di)
                for pi, part in enumerate(PARTICIPANTS):
                    sc(ws, r, cs+pi, fn(date, part), bg=bg, fmt=fmt, align="right")
                ws.cell(row=r, column=cs+np_).fill = _fill(bg)
            ws.row_dimensions[r].height = 14; r += 1
    ws.column_dimensions["A"].width = 38
    for di in range(nd):
        cs = dc(di)
        for pi in range(np_): ws.column_dimensions[get_column_letter(cs+pi)].width = 14
        ws.column_dimensions[get_column_letter(cs+np_)].width = 2

def build_dod_sheet(wb, dates, oi_row_map, tv_row_map):
    ws = wb.create_sheet("Day-on-Day Changes")
    ws.freeze_panes = "B4"
    if len(dates) < 2:
        ws.cell(1, 1).value = "Provide at least 2 dates for DoD analysis."
        return
    OI = "Raw OI Data"; TV = "Raw TV Data"
    def R(d,p,k): return ref(OI, oi_row_map, d, p, k)
    pairs = [(dates[i], dates[i+1]) for i in range(len(dates)-1)]
    np_ = len(PARTICIPANTS)
    def dc(di): return 2 + di * (np_ + 1)
    total_cols = 1 + len(pairs) * (np_ + 1)
    title_row(ws, 1, "NSE FAO — Day-on-Day OI & Volume Changes", total_cols)
    sc(ws, 2, 1, "Metric", bg=C["hdr_bg"], fg=C["hdr_fg"], bold=True, align="center")
    for di, (d0, d1) in enumerate(pairs):
        cs = dc(di)
        ws.merge_cells(start_row=2, start_column=cs, end_row=2, end_column=cs+np_-1)
        sc(ws, 2, cs, f"{d0} → {d1}", bg=C["hdr_bg"], fg=C["hdr_fg"],
           bold=True, align="center", wrap=True)
        for off in range(1, np_+1):
            ws.cell(row=2, column=cs+off).fill = _fill(C["hdr_bg"])
            ws.cell(row=2, column=cs+off).border = _bdr()
    sc(ws, 3, 1, "", bg=C["sect_bg"])
    for di in range(len(pairs)):
        cs = dc(di)
        for pi, part in enumerate(PARTICIPANTS):
            sc(ws, 3, cs+pi, part, bg=C["sect_bg"], fg=C["sect_fg"], bold=True, align="center")
        ws.cell(row=3, column=cs+np_).fill = _fill(C["sect_bg"])
    ws.row_dimensions[2].height = 28; ws.row_dimensions[3].height = 16

    dod_metrics = [
        ("OI CHANGES  (current session − previous session)", [
            ("Δ Fut Index Long",    lambda d0,d1,p: f"={R(d1,p,'FutIdxL')}-{R(d0,p,'FutIdxL')}", "#,##0;[Red]-#,##0;-"),
            ("Δ Fut Index Short",   lambda d0,d1,p: f"={R(d1,p,'FutIdxS')}-{R(d0,p,'FutIdxS')}", "#,##0;[Red]-#,##0;-"),
            ("Δ Net Fut Index OI",  lambda d0,d1,p: f"=({R(d1,p,'FutIdxL')}-{R(d1,p,'FutIdxS')})-({R(d0,p,'FutIdxL')}-{R(d0,p,'FutIdxS')})", "#,##0;[Red]-#,##0;-"),
            ("Δ Fut Stock Long",    lambda d0,d1,p: f"={R(d1,p,'FutStkL')}-{R(d0,p,'FutStkL')}", "#,##0;[Red]-#,##0;-"),
            ("Δ Fut Stock Short",   lambda d0,d1,p: f"={R(d1,p,'FutStkS')}-{R(d0,p,'FutStkS')}", "#,##0;[Red]-#,##0;-"),
            ("Δ Net Fut Stock OI",  lambda d0,d1,p: f"=({R(d1,p,'FutStkL')}-{R(d1,p,'FutStkS')})-({R(d0,p,'FutStkL')}-{R(d0,p,'FutStkS')})", "#,##0;[Red]-#,##0;-"),
            ("Δ Idx Call Long OI",  lambda d0,d1,p: f"={R(d1,p,'OptIdxCL')}-{R(d0,p,'OptIdxCL')}", "#,##0;[Red]-#,##0;-"),
            ("Δ Idx Put Long OI",   lambda d0,d1,p: f"={R(d1,p,'OptIdxPL')}-{R(d0,p,'OptIdxPL')}", "#,##0;[Red]-#,##0;-"),
            ("Δ Index OI PCR",      lambda d0,d1,p: f"=IFERROR({R(d1,p,'OptIdxPL')}/{R(d1,p,'OptIdxCL')},0)-IFERROR({R(d0,p,'OptIdxPL')}/{R(d0,p,'OptIdxCL')},0)", "0.00"),
            ("Δ Total Long OI",     lambda d0,d1,p: f"={R(d1,p,'TotL')}-{R(d0,p,'TotL')}", "#,##0;[Red]-#,##0;-"),
            ("Δ Total Short OI",    lambda d0,d1,p: f"={R(d1,p,'TotS')}-{R(d0,p,'TotS')}", "#,##0;[Red]-#,##0;-"),
            ("Δ Net Total OI",      lambda d0,d1,p: f"=({R(d1,p,'TotL')}-{R(d1,p,'TotS')})-({R(d0,p,'TotL')}-{R(d0,p,'TotS')})", "#,##0;[Red]-#,##0;-"),
        ]),
    ]
    if tv_row_map:
        def TV_(d,p,k): return ref(TV, tv_row_map, d, p, k)
        dod_metrics.append(("VOLUME CHANGES", [
            ("Δ Fut Idx Buy Vol",   lambda d0,d1,p: f"={TV_(d1,p,'FutIdxL')}-{TV_(d0,p,'FutIdxL')}", "#,##0;[Red]-#,##0;-"),
            ("Δ Idx Call Buy Vol",  lambda d0,d1,p: f"={TV_(d1,p,'OptIdxCL')}-{TV_(d0,p,'OptIdxCL')}", "#,##0;[Red]-#,##0;-"),
            ("Δ Idx Put Buy Vol",   lambda d0,d1,p: f"={TV_(d1,p,'OptIdxPL')}-{TV_(d0,p,'OptIdxPL')}", "#,##0;[Red]-#,##0;-"),
            ("Δ Vol PCR (Index)",   lambda d0,d1,p: f"=IFERROR({TV_(d1,p,'OptIdxPL')}/{TV_(d1,p,'OptIdxCL')},0)-IFERROR({TV_(d0,p,'OptIdxPL')}/{TV_(d0,p,'OptIdxCL')},0)", "0.00"),
            ("Δ Total Buy Vol",     lambda d0,d1,p: f"={TV_(d1,p,'TotL')}-{TV_(d0,p,'TotL')}", "#,##0;[Red]-#,##0;-"),
        ]))

    r = 4
    for group_label, metrics in dod_metrics:
        section_row(ws, r, group_label, total_cols); r += 1
        for mi, (label, fn, fmt) in enumerate(metrics):
            bg = C["alt"] if mi % 2 == 0 else C["white"]
            sc(ws, r, 1, label, bg=C["label_bg"], fg=C["label_fg"], align="left")
            for di, (d0, d1) in enumerate(pairs):
                cs = dc(di)
                for pi, part in enumerate(PARTICIPANTS):
                    sc(ws, r, cs+pi, fn(d0, d1, part), bg=bg, fmt=fmt, align="right")
                ws.cell(row=r, column=cs+np_).fill = _fill(bg)
            ws.row_dimensions[r].height = 14; r += 1
    ws.column_dimensions["A"].width = 28
    for di in range(len(pairs)):
        cs = dc(di)
        for pi in range(np_): ws.column_dimensions[get_column_letter(cs+pi)].width = 14
        ws.column_dimensions[get_column_letter(cs+np_)].width = 2

# ══════════════════════════════════════════════════════════════════════════════
# SHEET BUILDERS — OPTION CHAIN BASED
# ══════════════════════════════════════════════════════════════════════════════
def build_chain_view_sheet(wb, chain_dates, chain_df):
    """Strike × date grid: OI, ΔOI, Volume, IV, LTP, PCR per strike."""
    ws = wb.create_sheet("Option Chain View")
    ws.freeze_panes = "C4"
    nd = len(chain_dates)

    # Columns per date: C_OI, C_ΔOI, C_VOL, C_IV, | Strike | P_OI, P_ΔOI, P_VOL, P_IV, PCR
    CALL_FIELDS = [("C_OI","Call OI"), ("C_CHNG_OI","Call ΔOI"),
                   ("C_VOLUME","Call Vol"), ("C_IV","Call IV"), ("C_LTP","Call LTP")]
    PUT_FIELDS  = [("P_OI","Put OI"),  ("P_CHNG_OI","Put ΔOI"),
                   ("P_VOLUME","Put Vol"),  ("P_IV","Put IV"),  ("P_LTP","Put LTP")]
    STRIKE_COL  = 1   # Strike always column A
    # Layout: Strike (A) | date0: 5 call cols, 5 put cols, PCR | date1: ... 
    COLS_PER_DATE = len(CALL_FIELDS) + len(PUT_FIELDS) + 1  # +1 for PCR
    SEP = 1  # separator column between dates

    def date_call_start(di): return 2 + di * (COLS_PER_DATE + SEP)
    def date_put_start(di): return date_call_start(di) + len(CALL_FIELDS)
    def date_pcr_col(di): return date_put_start(di) + len(PUT_FIELDS)

    total_cols = 1 + nd * (COLS_PER_DATE + SEP)
    title_row(ws, 1, "NSE FAO — Option Chain View (Strike-wise, All Dates)", total_cols)
    ws.row_dimensions[1].height = 22

    # Row 2: date spans with call/put super-headers
    sc(ws, 2, 1, "Strike", bg=C["hdr_bg"], fg=C["hdr_fg"], bold=True, align="center")
    for di, date in enumerate(chain_dates):
        cs_c = date_call_start(di)
        cs_p = date_put_start(di)
        cs_pcr = date_pcr_col(di)
        # Date span header
        ws.merge_cells(start_row=2, start_column=cs_c,
                       end_row=2, end_column=cs_pcr)
        sc(ws, 2, cs_c, date, bg=C["hdr_bg"], fg=C["hdr_fg"], bold=True, align="center")
        for off in range(1, COLS_PER_DATE + SEP):
            ws.cell(row=2, column=cs_c+off).fill = _fill(C["hdr_bg"])
            ws.cell(row=2, column=cs_c+off).border = _bdr()

    # Row 3: CALLS / PUTS section labels
    sc(ws, 3, 1, "", bg=C["sub_bg"])
    for di in range(nd):
        cs_c = date_call_start(di)
        cs_p = date_put_start(di)
        cs_pcr = date_pcr_col(di)
        ws.merge_cells(start_row=3, start_column=cs_c,
                       end_row=3, end_column=cs_c + len(CALL_FIELDS) - 1)
        sc(ws, 3, cs_c, "← CALLS", bg=C["call_bg"], fg=C["blue_fg"], bold=True, align="center")
        for off in range(1, len(CALL_FIELDS)):
            ws.cell(row=3, column=cs_c+off).fill = _fill(C["call_bg"])
            ws.cell(row=3, column=cs_c+off).border = _bdr()
        ws.merge_cells(start_row=3, start_column=cs_p,
                       end_row=3, end_column=cs_p + len(PUT_FIELDS) - 1)
        sc(ws, 3, cs_p, "PUTS →", bg=C["put_bg"], fg=C["amber_fg"], bold=True, align="center")
        for off in range(1, len(PUT_FIELDS)):
            ws.cell(row=3, column=cs_p+off).fill = _fill(C["put_bg"])
            ws.cell(row=3, column=cs_p+off).border = _bdr()
        sc(ws, 3, cs_pcr, "PCR", bg=C["purple_bg"], fg=C["purple_fg"], bold=True, align="center")

    # Row 4: field headers
    sc(ws, 4, 1, "Strike", bg=C["sub_bg"], fg=C["hdr_fg"], bold=True, align="center")
    for di in range(nd):
        cs_c = date_call_start(di)
        cs_p = date_put_start(di)
        cs_pcr = date_pcr_col(di)
        for ci, (_, lbl) in enumerate(CALL_FIELDS):
            sc(ws, 4, cs_c+ci, lbl, bg=C["call_bg"], fg=C["blue_fg"], bold=True, align="center", wrap=True)
        for pi, (_, lbl) in enumerate(PUT_FIELDS):
            sc(ws, 4, cs_p+pi, lbl, bg=C["put_bg"], fg=C["amber_fg"], bold=True, align="center", wrap=True)
        sc(ws, 4, cs_pcr, "P/C OI", bg=C["purple_bg"], fg=C["purple_fg"], bold=True, align="center")

    ws.row_dimensions[2].height = 16
    ws.row_dimensions[3].height = 14
    ws.row_dimensions[4].height = 28

    # Get common strike universe
    all_strikes = sorted(chain_df["Strike"].unique())
    # ATM for latest date
    latest = chain_df[chain_df["Date"] == chain_dates[-1]]
    atm = compute_atm(latest) if not latest.empty else None

    r = 5
    for strike in all_strikes:
        is_atm = (strike == atm)
        bg_row = C["wall_bg"] if is_atm else (C["alt"] if r % 2 == 0 else C["white"])
        sc(ws, r, 1, strike, bg=bg_row, bold=is_atm, align="right", fmt="#,##0")
        for di, date in enumerate(chain_dates):
            sub = chain_df[(chain_df["Date"] == date) & (chain_df["Strike"] == strike)]
            cs_c = date_call_start(di); cs_p = date_put_start(di); cs_pcr = date_pcr_col(di)
            if sub.empty:
                for ci in range(len(CALL_FIELDS)): sc(ws, r, cs_c+ci, None, bg=bg_row)
                for pi in range(len(PUT_FIELDS)):  sc(ws, r, cs_p+pi, None, bg=bg_row)
                sc(ws, r, cs_pcr, None, bg=bg_row)
            else:
                row_data = sub.iloc[0]
                call_fmts = ["#,##0","#,##0","#,##0","0.00","0.00"]
                put_fmts  = ["#,##0","#,##0","#,##0","0.00","0.00"]
                for ci, (col, _) in enumerate(CALL_FIELDS):
                    v = row_data.get(col)
                    sc(ws, r, cs_c+ci,
                       (float(v) if v is not None and not pd.isna(v) else None),
                       bg=C["call_bg"] if not is_atm else C["wall_bg"],
                       fmt=call_fmts[ci], align="right")
                for pi, (col, _) in enumerate(PUT_FIELDS):
                    v = row_data.get(col)
                    sc(ws, r, cs_p+pi,
                       (float(v) if v is not None and not pd.isna(v) else None),
                       bg=C["put_bg"] if not is_atm else C["wall_bg"],
                       fmt=put_fmts[pi], align="right")
                c_oi = row_data.get("C_OI") or 0
                p_oi = row_data.get("P_OI") or 0
                pcr_val = round(p_oi / c_oi, 2) if c_oi else None
                sc(ws, r, cs_pcr, pcr_val, bg=C["purple_bg"] if not is_atm else C["wall_bg"],
                   fg=C["purple_fg"], align="right", fmt="0.00")
        ws.row_dimensions[r].height = 13
        r += 1

    ws.column_dimensions["A"].width = 10
    for di in range(nd):
        cs_c = date_call_start(di)
        cs_p = date_put_start(di)
        cs_pcr = date_pcr_col(di)
        for ci in range(len(CALL_FIELDS)):
            ws.column_dimensions[get_column_letter(cs_c+ci)].width = 12
        for pi in range(len(PUT_FIELDS)):
            ws.column_dimensions[get_column_letter(cs_p+pi)].width = 12
        ws.column_dimensions[get_column_letter(cs_pcr)].width = 7
        # separator
        sep_col = cs_pcr + 1
        ws.column_dimensions[get_column_letter(sep_col)].width = 2

def build_max_pain_sheet(wb, chain_dates, chain_df, chain_analytics):
    ws = wb.create_sheet("Max Pain")
    ws.freeze_panes = "B3"
    nd = len(chain_dates)
    all_strikes = sorted(chain_df["Strike"].unique())

    # Summary at top
    SUMM_COLS = 1 + nd
    title_row(ws, 1, "NSE FAO — Max Pain Analysis (Aggregate OI Pain by Strike)", SUMM_COLS + 10)

    # Summary header row
    sc(ws, 2, 1, "Summary", bg=C["sect_bg"], fg=C["sect_fg"], bold=True, align="left")
    for di, date in enumerate(chain_dates):
        sc(ws, 2, 2+di, date, bg=C["sect_bg"], fg=C["sect_fg"], bold=True, align="center", wrap=True)
    ws.row_dimensions[2].height = 28

    summary_rows = [
        ("Max Pain Strike",  "max_pain",    "#,##0"),
        ("ATM Strike",       "atm",         "#,##0"),
        ("Aggregate OI PCR", "oi_pcr",      "0.000"),
        ("Total Call OI",    "total_call_oi","#,##0"),
        ("Total Put OI",     "total_put_oi", "#,##0"),
    ]
    for ri, (label, key, fmt) in enumerate(summary_rows):
        bg = C["alt"] if ri % 2 == 0 else C["white"]
        bold = (label == "Max Pain Strike")
        bg_val = C["amber_bg"] if label == "Max Pain Strike" else bg
        sc(ws, 3+ri, 1, label, bg=C["label_bg"], fg=C["label_fg"],
           bold=bold, align="left")
        for di, date in enumerate(chain_dates):
            val = chain_analytics.get(date, {}).get(key)
            sc(ws, 3+ri, 2+di, val, bg=bg_val,
               fg=C["amber_fg"] if label == "Max Pain Strike" else "000000",
               bold=bold, align="right", fmt=fmt)
        ws.row_dimensions[3+ri].height = 14

    # Blank row
    blank_r = 3 + len(summary_rows) + 1
    ws.row_dimensions[blank_r].height = 8

    # Per-strike pain table
    detail_r = blank_r + 1
    section_row(ws, detail_r, "Per-Strike Pain Values by Date  (lower = more likely to be Max Pain)", SUMM_COLS + 10)
    detail_r += 1
    ws.row_dimensions[detail_r].height = 28

    # Header: Strike | date0 pain | date1 pain | ...
    sc(ws, detail_r, 1, "Strike", bg=C["sub_bg"], fg=C["hdr_fg"], bold=True, align="center")
    for di, date in enumerate(chain_dates):
        sc(ws, detail_r, 2+di, date, bg=C["sub_bg"], fg=C["hdr_fg"],
           bold=True, align="center", wrap=True)

    detail_r += 1
    for strike in all_strikes:
        bg = C["alt"] if detail_r % 2 == 0 else C["white"]
        sc(ws, detail_r, 1, strike, bg=bg, align="right", fmt="#,##0")
        for di, date in enumerate(chain_dates):
            pain_map = chain_analytics.get(date, {}).get("pain_map", {})
            val = pain_map.get(strike)
            is_mp = (val is not None and
                     val == min((v for v in pain_map.values() if v is not None), default=None))
            sc(ws, detail_r, 2+di,
               (int(val) if val is not None else None),
               bg=C["amber_bg"] if is_mp else bg,
               fg=C["amber_fg"] if is_mp else "000000",
               bold=is_mp, align="right", fmt="#,##0")
        ws.row_dimensions[detail_r].height = 13
        detail_r += 1

    ws.column_dimensions["A"].width = 10
    for di in range(nd):
        ws.column_dimensions[get_column_letter(2+di)].width = 18

def build_oi_walls_sheet(wb, chain_dates, chain_analytics, top_n=10):
    ws = wb.create_sheet("OI Walls & Support")
    ws.freeze_panes = "A3"
    nd = len(chain_dates)
    # Layout: for each date, a block: Call Wall table | Put Wall table, side by side
    WALL_COLS  = ["Strike", "Call OI", "Δ Call OI", "Call Vol", "% of Total Call OI"]
    WALL_COLS_P= ["Strike", "Put OI",  "Δ Put OI",  "Put Vol",  "% of Total Put OI"]
    BLOCK_W    = len(WALL_COLS) + len(WALL_COLS_P) + 1  # +1 separator

    total_cols = nd * BLOCK_W
    title_row(ws, 1, f"NSE FAO — OI Walls: Call Resistance & Put Support (Top {top_n} Strikes)", total_cols)

    r_base = 2
    for di, date in enumerate(chain_dates):
        col_start = 1 + di * BLOCK_W
        walls = chain_analytics.get(date, {})
        call_walls = walls.get("call_walls", pd.DataFrame())
        put_walls  = walls.get("put_walls",  pd.DataFrame())
        total_c_oi = walls.get("total_call_oi", 1) or 1
        total_p_oi = walls.get("total_put_oi",  1) or 1

        # Date header spanning the block
        ws.merge_cells(start_row=r_base, start_column=col_start,
                       end_row=r_base, end_column=col_start + BLOCK_W - 2)
        sc(ws, r_base, col_start, date, bg=C["hdr_bg"], fg=C["hdr_fg"],
           bold=True, align="center")
        for off in range(1, BLOCK_W - 1):
            ws.cell(row=r_base, column=col_start+off).fill = _fill(C["hdr_bg"])
            ws.cell(row=r_base, column=col_start+off).border = _bdr()

        # Call wall header
        ws.merge_cells(start_row=r_base+1, start_column=col_start,
                       end_row=r_base+1, end_column=col_start+len(WALL_COLS)-1)
        sc(ws, r_base+1, col_start, "⬆ CALL WALL (Resistance)", bg=C["call_bg"],
           fg=C["blue_fg"], bold=True, align="center")
        for off in range(1, len(WALL_COLS)):
            ws.cell(row=r_base+1, column=col_start+off).fill = _fill(C["call_bg"])
            ws.cell(row=r_base+1, column=col_start+off).border = _bdr()

        # Put wall header
        put_start = col_start + len(WALL_COLS)
        ws.merge_cells(start_row=r_base+1, start_column=put_start,
                       end_row=r_base+1, end_column=put_start+len(WALL_COLS_P)-1)
        sc(ws, r_base+1, put_start, "⬇ PUT WALL (Support)", bg=C["put_bg"],
           fg=C["amber_fg"], bold=True, align="center")
        for off in range(1, len(WALL_COLS_P)):
            ws.cell(row=r_base+1, column=put_start+off).fill = _fill(C["put_bg"])
            ws.cell(row=r_base+1, column=put_start+off).border = _bdr()

        # Column headers
        for ci, h in enumerate(WALL_COLS):
            sc(ws, r_base+2, col_start+ci, h, bg=C["call_bg"], fg=C["blue_fg"],
               bold=True, align="center", wrap=True)
        for ci, h in enumerate(WALL_COLS_P):
            sc(ws, r_base+2, put_start+ci, h, bg=C["put_bg"], fg=C["amber_fg"],
               bold=True, align="center", wrap=True)
        ws.row_dimensions[r_base+2].height = 28

        # Data rows
        for ri in range(top_n):
            row_r = r_base + 3 + ri
            bg = C["alt"] if ri % 2 == 0 else C["white"]
            ws.row_dimensions[row_r].height = 14
            # Call side
            if ri < len(call_walls):
                cw = call_walls.iloc[ri]
                sc(ws, row_r, col_start,   cw["Strike"],    bg=C["call_bg"], align="right", fmt="#,##0")
                sc(ws, row_r, col_start+1, cw["C_OI"],      bg=C["call_bg"], align="right", fmt="#,##0")
                sc(ws, row_r, col_start+2, cw["C_CHNG_OI"], bg=C["call_bg"], align="right", fmt="#,##0;[Red]-#,##0;-")
                sc(ws, row_r, col_start+3, cw["C_VOLUME"],  bg=C["call_bg"], align="right", fmt="#,##0")
                pct = round(cw["C_OI"] / total_c_oi, 4) if total_c_oi else None
                sc(ws, row_r, col_start+4, pct, bg=C["call_bg"], align="right", fmt="0.0%")
            else:
                for ci in range(len(WALL_COLS)):
                    sc(ws, row_r, col_start+ci, None, bg=C["call_bg"])
            # Put side
            if ri < len(put_walls):
                pw = put_walls.iloc[ri]
                sc(ws, row_r, put_start,   pw["Strike"],    bg=C["put_bg"], align="right", fmt="#,##0")
                sc(ws, row_r, put_start+1, pw["P_OI"],      bg=C["put_bg"], align="right", fmt="#,##0")
                sc(ws, row_r, put_start+2, pw["P_CHNG_OI"], bg=C["put_bg"], align="right", fmt="#,##0;[Red]-#,##0;-")
                sc(ws, row_r, put_start+3, pw["P_VOLUME"],  bg=C["put_bg"], align="right", fmt="#,##0")
                pct = round(pw["P_OI"] / total_p_oi, 4) if total_p_oi else None
                sc(ws, row_r, put_start+4, pct, bg=C["put_bg"], align="right", fmt="0.0%")
            else:
                for ci in range(len(WALL_COLS_P)):
                    sc(ws, row_r, put_start+ci, None, bg=C["put_bg"])

    ws.row_dimensions[r_base].height   = 18
    ws.row_dimensions[r_base+1].height = 16
    for di in range(nd):
        col_start = 1 + di * BLOCK_W
        for ci in range(BLOCK_W - 1):
            ws.column_dimensions[get_column_letter(col_start+ci)].width = 14
        ws.column_dimensions[get_column_letter(col_start + BLOCK_W - 1)].width = 2

def build_iv_skew_sheet(wb, chain_dates, chain_df, chain_analytics):
    ws = wb.create_sheet("IV Skew")
    ws.freeze_panes = "B4"
    nd = len(chain_dates)
    all_strikes = sorted(chain_df["Strike"].unique())

    COLS_PER_DATE = 3 + 1  # C_IV, P_IV, IV_SKEW + sep
    total_cols = 1 + nd * COLS_PER_DATE
    title_row(ws, 1, "NSE FAO — Implied Volatility Skew by Strike", total_cols)

    sc(ws, 2, 1, "Strike", bg=C["hdr_bg"], fg=C["hdr_fg"], bold=True, align="center")
    for di, date in enumerate(chain_dates):
        cs = 2 + di * COLS_PER_DATE
        ws.merge_cells(start_row=2, start_column=cs, end_row=2, end_column=cs+2)
        sc(ws, 2, cs, date, bg=C["hdr_bg"], fg=C["hdr_fg"], bold=True, align="center")
        for off in range(1, 4):
            ws.cell(row=2, column=cs+off).fill = _fill(C["hdr_bg"])
            ws.cell(row=2, column=cs+off).border = _bdr()

    hdrs = ["Call IV", "Put IV", "Put−Call Skew"]
    sc(ws, 3, 1, "", bg=C["sub_bg"])
    for di in range(nd):
        cs = 2 + di * COLS_PER_DATE
        bgs = [C["call_bg"], C["put_bg"], C["purple_bg"]]
        fgs = [C["blue_fg"], C["amber_fg"], C["purple_fg"]]
        for ci, (h, bg_, fg_) in enumerate(zip(hdrs, bgs, fgs)):
            sc(ws, 3, cs+ci, h, bg=bg_, fg=fg_, bold=True, align="center", wrap=True)
        ws.cell(row=3, column=cs+3).fill = _fill(C["sub_bg"])
    ws.row_dimensions[2].height = 16; ws.row_dimensions[3].height = 28

    latest_date = chain_dates[-1]
    latest = chain_df[chain_df["Date"] == latest_date]
    atm = compute_atm(latest) if not latest.empty else None

    r = 4
    for strike in all_strikes:
        is_atm = (strike == atm)
        bg_base = C["wall_bg"] if is_atm else (C["alt"] if r % 2 == 0 else C["white"])
        sc(ws, r, 1, strike, bg=bg_base, bold=is_atm, align="right", fmt="#,##0")
        for di, date in enumerate(chain_dates):
            cs = 2 + di * COLS_PER_DATE
            sub = chain_df[(chain_df["Date"] == date) & (chain_df["Strike"] == strike)]
            if sub.empty:
                for ci in range(3): sc(ws, r, cs+ci, None, bg=bg_base)
            else:
                row_data = sub.iloc[0]
                c_iv = row_data.get("C_IV")
                p_iv = row_data.get("P_IV")
                skew = (round(p_iv - c_iv, 2)
                        if (c_iv is not None and p_iv is not None
                            and not pd.isna(c_iv) and not pd.isna(p_iv))
                        else None)
                c_iv_v = float(c_iv) if (c_iv is not None and not pd.isna(c_iv)) else None
                p_iv_v = float(p_iv) if (p_iv is not None and not pd.isna(p_iv)) else None
                sc(ws, r, cs,   c_iv_v, bg=C["call_bg"] if not is_atm else bg_base,
                   fmt="0.00", align="right")
                sc(ws, r, cs+1, p_iv_v, bg=C["put_bg"] if not is_atm else bg_base,
                   fmt="0.00", align="right")
                skew_bg = (C["red_bg"] if (skew and skew > 2)
                           else C["green_bg"] if (skew and skew < -2)
                           else (C["purple_bg"] if not is_atm else bg_base))
                skew_fg = (C["red_fg"] if (skew and skew > 2)
                           else C["green_fg"] if (skew and skew < -2)
                           else C["purple_fg"])
                sc(ws, r, cs+2, skew, bg=skew_bg, fg=skew_fg,
                   fmt="0.00;[Red]-0.00", align="right")
            ws.cell(row=r, column=cs+3).fill = _fill(bg_base)
        ws.row_dimensions[r].height = 13; r += 1

    ws.column_dimensions["A"].width = 10
    for di in range(nd):
        cs = 2 + di * COLS_PER_DATE
        ws.column_dimensions[get_column_letter(cs)].width   = 10
        ws.column_dimensions[get_column_letter(cs+1)].width = 10
        ws.column_dimensions[get_column_letter(cs+2)].width = 14
        ws.column_dimensions[get_column_letter(cs+3)].width = 2

def build_sentiment_sheet(wb, dates, oi_row_map, tv_row_map, chain_analytics):
    ws = wb.create_sheet("Sentiment Summary")
    ws.freeze_panes = "C3"
    OI = "Raw OI Data"
    nd = len(dates)
    n_cols = 2 + nd + 2  # label + signal | dates | latest | note
    title_row(ws, 1, "NSE FAO — Composite Sentiment Summary", n_cols + 2)

    hdrs = ["Participant", "Signal"] + dates + ["Latest Signal", "Interpretation"]
    for ci, h in enumerate(hdrs, 1):
        sc(ws, 2, ci, h, bg=C["sect_bg"], fg=C["sect_fg"],
           bold=True, align="center", wrap=True)
    ws.row_dimensions[2].height = 28

    def R(d, p, k): return ref(OI, oi_row_map, d, p, k)

    def sentiment_formula(d, p):
        net_fi = f"({R(d,p,'FutIdxL')}-{R(d,p,'FutIdxS')})"
        net_si = f"({R(d,p,'FutStkL')}-{R(d,p,'FutStkS')})"
        pcr    = f"IFERROR({R(d,p,'OptIdxPL')}/{R(d,p,'OptIdxCL')},1)"
        ls_tot = f"IFERROR({R(d,p,'TotL')}/{R(d,p,'TotS')},1)"
        return (f'=IF(AND({net_fi}>0,{pcr}<1,{ls_tot}>1),"🟢 Bullish",'
                f'IF(AND({net_fi}<0,{pcr}>1.5,{ls_tot}<1),"🔴 Bearish",'
                f'IF(AND({net_fi}<0,{pcr}>1.2),"🟠 Cautious",'
                f'IF(AND({net_fi}>0,{pcr}<0.8),"🟢 Bullish","🟡 Neutral"))))')

    part_signals = [
        ("Fut Index Net OI",  lambda d,p: f'=IF({R(d,p,"FutIdxL")}-{R(d,p,"FutIdxS")}>0,"▲ Net Long","▼ Net Short")', None),
        ("Fut Stock Net OI",  lambda d,p: f'=IF({R(d,p,"FutStkL")}-{R(d,p,"FutStkS")}>0,"▲ Net Long","▼ Net Short")', None),
        ("Index OI PCR",      lambda d,p: f'=IFERROR({R(d,p,"OptIdxPL")}/{R(d,p,"OptIdxCL")},0)', "0.00"),
        ("Total L/S Ratio",   lambda d,p: f'=IFERROR({R(d,p,"TotL")}/{R(d,p,"TotS")},0)', "0.00x"),
        ("Composite Signal",  sentiment_formula, None),
    ]

    INTERPRETATIONS = {
        "Client":  "Retail — usually contra-indicator; track when extreme",
        "DII":     "Domestic Institutions — typically hedged; PCR key",
        "FII":     "Foreign Institutions — directional alpha signal",
        "Pro":     "Proprietary desks — fast money; mean-reverting",
    }

    r = 3
    for part in PARTICIPANTS:
        section_row(ws, r, part, n_cols + 2); r += 1
        for si, (sig_label, sig_fn, fmt) in enumerate(part_signals):
            is_composite = (sig_label == "Composite Signal")
            bg = C["alt"] if si % 2 == 0 else C["white"]
            cell_bg = C["blue_bg"] if is_composite else bg
            cell_fg = C["blue_fg"] if is_composite else "000000"
            sc(ws, r, 1, "", bg=bg)
            sc(ws, r, 2, sig_label, bg=C["label_bg"], fg=C["label_fg"],
               bold=is_composite, align="left")
            date_cells = []
            for di, date in enumerate(dates):
                formula = sig_fn(date, part)
                c = sc(ws, r, 3+di, formula, bg=cell_bg, fg=cell_fg,
                       bold=is_composite, align="center", fmt=fmt)
                date_cells.append(f"{get_column_letter(3+di)}{r}")
            # Latest signal = last date cell
            last_cell = date_cells[-1]
            sc(ws, r, 3+nd, f"={last_cell}", bg=C["amber_bg"] if is_composite else cell_bg,
               fg=C["amber_fg"] if is_composite else cell_fg,
               bold=is_composite, align="center", fmt=fmt)
            # Interpretation only on composite row
            interp = INTERPRETATIONS.get(part, "") if is_composite else ""
            sc(ws, r, 3+nd+1, interp, bg=bg, fg="6B7280",
               italic=True, align="left", wrap=True)
            ws.row_dimensions[r].height = 14; r += 1

    # Chain-level summary block
    if chain_analytics:
        section_row(ws, r, "CHAIN-LEVEL SIGNALS  (Market-wide, all participants)", n_cols + 2)
        r += 1
        chain_signal_rows = [
            ("Max Pain Strike",    "max_pain",  "#,##0"),
            ("ATM Strike",         "atm",       "#,##0"),
            ("Aggregate OI PCR",   "oi_pcr",    "0.000"),
            ("Aggregate Volume PCR","vol_pcr",   "0.000"),
        ]
        for mi, (label, key, fmt) in enumerate(chain_signal_rows):
            bg = C["alt"] if mi % 2 == 0 else C["white"]
            sc(ws, r, 1, "", bg=bg)
            sc(ws, r, 2, label, bg=C["label_bg"], fg=C["label_fg"], align="left")
            for di, date in enumerate(dates):
                val = chain_analytics.get(date, {}).get(key)
                sc(ws, r, 3+di, val, bg=C["amber_bg"] if "Pain" in label else bg,
                   fg=C["amber_fg"] if "Pain" in label else "000000",
                   bold="Pain" in label, align="center", fmt=fmt)
            # Latest (use last date)
            last_val = chain_analytics.get(dates[-1], {}).get(key) if dates else None
            sc(ws, r, 3+nd, last_val,
               bg=C["amber_bg"] if "Pain" in label else bg,
               fg=C["amber_fg"] if "Pain" in label else "000000",
               bold="Pain" in label, align="center", fmt=fmt)
            sc(ws, r, 3+nd+1, None, bg=bg)
            ws.row_dimensions[r].height = 14; r += 1

    col_widths = [2, 22] + [16]*nd + [16, 36]
    for ci, w in enumerate(col_widths, 1):
        if ci <= n_cols + 2:
            ws.column_dimensions[get_column_letter(ci)].width = w

# ══════════════════════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════════════════════
# ══════════════════════════════════════════════════════════════════════════════
# JSON EXPORT  (data contract for the website)
# ══════════════════════════════════════════════════════════════════════════════
def _safe(v):
    """Convert numpy/pandas scalars to plain Python for JSON serialisation."""
    if v is None:
        return None
    try:
        import math
        if isinstance(v, float) and math.isnan(v):
            return None
    except Exception:
        pass
    if hasattr(v, "item"):          # numpy scalar
        return v.item()
    if hasattr(v, "tolist"):        # numpy array
        return v.tolist()
    return v


def _participant_record(row: pd.Series) -> dict:
    """Build a full participant record from a raw OI or TV dataframe row."""
    d = {
        "fut_idx_long":   int(_safe(row.get("Future Index Long",  0)) or 0),
        "fut_idx_short":  int(_safe(row.get("Future Index Short", 0)) or 0),
        "fut_stk_long":   int(_safe(row.get("Future Stock Long",  0)) or 0),
        "fut_stk_short":  int(_safe(row.get("Future Stock Short", 0)) or 0),
        "opt_idx_cl":     int(_safe(row.get("Option Index Call Long",  0)) or 0),
        "opt_idx_pl":     int(_safe(row.get("Option Index Put Long",   0)) or 0),
        "opt_idx_cs":     int(_safe(row.get("Option Index Call Short", 0)) or 0),
        "opt_idx_ps":     int(_safe(row.get("Option Index Put Short",  0)) or 0),
        "opt_stk_cl":     int(_safe(row.get("Option Stock Call Long",  0)) or 0),
        "opt_stk_pl":     int(_safe(row.get("Option Stock Put Long",   0)) or 0),
        "opt_stk_cs":     int(_safe(row.get("Option Stock Call Short", 0)) or 0),
        "opt_stk_ps":     int(_safe(row.get("Option Stock Put Short",  0)) or 0),
        "total_long":     int(_safe(row.get("Total Long Contracts",    0)) or 0),
        "total_short":    int(_safe(row.get("Total Short Contracts",   0)) or 0),
    }
    # Derived fields
    d["net_fut_idx"]    = d["fut_idx_long"]  - d["fut_idx_short"]
    d["net_fut_stk"]    = d["fut_stk_long"]  - d["fut_stk_short"]
    d["net_futures"]    = d["net_fut_idx"]   + d["net_fut_stk"]
    d["net_idx_call"]   = d["opt_idx_cl"]    - d["opt_idx_cs"]
    d["net_idx_put"]    = d["opt_idx_pl"]    - d["opt_idx_ps"]
    d["net_stk_call"]   = d["opt_stk_cl"]    - d["opt_stk_cs"]
    d["net_stk_put"]    = d["opt_stk_pl"]    - d["opt_stk_ps"]
    d["net_options"]    = (d["net_idx_call"] + d["net_idx_put"] +
                           d["net_stk_call"] + d["net_stk_put"])
    d["net_total"]      = d["total_long"]    - d["total_short"]
    d["ls_ratio"]       = round(d["total_long"] / d["total_short"], 4) if d["total_short"] else None
    d["ls_fut_idx"]     = round(d["fut_idx_long"] / d["fut_idx_short"], 4) if d["fut_idx_short"] else None
    d["ls_fut_stk"]     = round(d["fut_stk_long"] / d["fut_stk_short"], 4) if d["fut_stk_short"] else None
    d["idx_oi_pcr"]     = round(d["opt_idx_pl"] / d["opt_idx_cl"], 4) if d["opt_idx_cl"] else None
    d["stk_oi_pcr"]     = round(d["opt_stk_pl"] / d["opt_stk_cl"], 4) if d["opt_stk_cl"] else None
    d["comb_oi_pcr"]    = (round((d["opt_idx_pl"] + d["opt_stk_pl"]) /
                                 (d["opt_idx_cl"] + d["opt_stk_cl"]), 4)
                           if (d["opt_idx_cl"] + d["opt_stk_cl"]) else None)
    d["net_idx_pcr"]    = (round((d["opt_idx_pl"] - d["opt_idx_ps"]) /
                                 (d["opt_idx_cl"] - d["opt_idx_cs"]), 4)
                           if (d["opt_idx_cl"] - d["opt_idx_cs"]) != 0 else None)
    d["short_pcr"]      = (round(d["opt_idx_ps"] / d["opt_idx_cs"], 4)
                           if d["opt_idx_cs"] else None)
    return d


def _sentiment_signal(rec: dict) -> str:
    """Derive composite sentiment string from a participant OI record."""
    net_fi  = rec.get("net_fut_idx", 0) or 0
    pcr     = rec.get("idx_oi_pcr")
    ls      = rec.get("ls_ratio")
    if pcr is None or ls is None:
        return "Neutral"
    if net_fi > 0 and pcr < 0.8  and ls > 1:  return "Bullish"
    if net_fi > 0 and pcr < 1.0  and ls > 1:  return "Bullish"
    if net_fi < 0 and pcr > 1.5  and ls < 1:  return "Bearish"
    if net_fi < 0 and pcr > 1.2:               return "Cautious"
    return "Neutral"


def _build_participant_date_block(df: pd.DataFrame, date: str) -> dict:
    """Return {participant: record} for a single date."""
    block = {}
    sub = df[df["Date"] == date]
    totals_long = totals_short = 0
    for part in PARTICIPANTS:
        row = sub[sub["Client Type"] == part]
        if row.empty:
            rec = _participant_record(pd.Series({c: 0 for c in PART_COLS}))
        else:
            rec = _participant_record(row.iloc[0])
        block[part] = rec
        totals_long  += rec["total_long"]
        totals_short += rec["total_short"]
    # market share
    for part in PARTICIPANTS:
        rec = block[part]
        rec["mkt_share_long"]  = round(rec["total_long"]  / totals_long,  4) if totals_long  else None
        rec["mkt_share_short"] = round(rec["total_short"] / totals_short, 4) if totals_short else None
        # market share by fut index long
        fut_idx_long_total = sum(block[p]["fut_idx_long"] for p in PARTICIPANTS)
        fut_stk_short_total = sum(block[p]["fut_stk_short"] for p in PARTICIPANTS)
        rec["mkt_share_fut_idx_long"]  = (round(rec["fut_idx_long"]  / fut_idx_long_total,  4)
                                           if fut_idx_long_total  else None)
        rec["mkt_share_fut_stk_short"] = (round(rec["fut_stk_short"] / fut_stk_short_total, 4)
                                           if fut_stk_short_total else None)
    block["_totals"] = {"total_long": totals_long, "total_short": totals_short}
    return block


def _dod_record(prev: dict, curr: dict) -> dict:
    """Day-on-day delta between two participant records."""
    skip = {"ls_ratio", "ls_fut_idx", "ls_fut_stk", "idx_oi_pcr", "stk_oi_pcr",
            "comb_oi_pcr", "net_idx_pcr", "short_pcr",
            "mkt_share_long", "mkt_share_short",
            "mkt_share_fut_idx_long", "mkt_share_fut_stk_short"}
    d = {}
    for k, v in curr.items():
        if k in skip:
            d[f"d_{k}"] = None
        elif isinstance(v, (int, float)) and v is not None:
            pv = prev.get(k)
            d[f"d_{k}"] = (v - pv) if (pv is not None and isinstance(pv, (int, float))) else None
        else:
            d[f"d_{k}"] = None
    # PCR deltas specifically
    for pcr_key in ("idx_oi_pcr", "stk_oi_pcr", "comb_oi_pcr"):
        cv = curr.get(pcr_key)
        pv = prev.get(pcr_key)
        d[f"d_{pcr_key}"] = (round(cv - pv, 4)
                              if (cv is not None and pv is not None) else None)
    return d


def _efficiency_record(oi_rec: dict, tv_rec: dict) -> dict:
    """OI/TV efficiency ratios for a participant."""
    pairs = [
        ("fut_idx", "fut_idx_long"),
        ("fut_stk", "fut_stk_long"),
        ("idx_call", "opt_idx_cl"),
        ("idx_put",  "opt_idx_pl"),
        ("stk_call", "opt_stk_cl"),
        ("stk_put",  "opt_stk_pl"),
        ("total",    "total_long"),
    ]
    d = {}
    for label, key in pairs:
        oi_v = oi_rec.get(key, 0) or 0
        tv_v = tv_rec.get(key, 0) or 0
        d[f"oi_tv_{label}"] = round(oi_v / tv_v, 4) if tv_v else None
    # net conviction pairs
    for seg, oi_l, oi_s, tv_l, tv_s in [
        ("fut_idx", "fut_idx_long", "fut_idx_short", "fut_idx_long", "fut_idx_short"),
        ("idx_call","opt_idx_cl",   "opt_idx_cs",    "opt_idx_cl",  "opt_idx_cs"),
        ("idx_put", "opt_idx_pl",   "opt_idx_ps",    "opt_idx_pl",  "opt_idx_ps"),
    ]:
        d[f"net_oi_{seg}"]  = (oi_rec.get(oi_l, 0) or 0) - (oi_rec.get(oi_s, 0) or 0)
        d[f"net_tv_{seg}"]  = (tv_rec.get(tv_l, 0) or 0) - (tv_rec.get(tv_s, 0) or 0)
    return d


def export_json(
    out_path: str,
    date_tag: str,
    excel_filename: str,
    oi_dates: list,
    oi_df: pd.DataFrame,
    tv_df,           # pd.DataFrame or None
    chain_df,        # pd.DataFrame or None
    chain_analytics: dict,
):
    """Build and write docs/data.json — the full data contract for the website."""

    # ── OI block ──────────────────────────────────────────────────────────
    oi_block = {"dates": oi_dates, "data": {}}
    for date in oi_dates:
        oi_block["data"][date] = _build_participant_date_block(oi_df, date)

    # ── TV block ──────────────────────────────────────────────────────────
    tv_block = {"dates": [], "data": {}}
    if tv_df is not None:
        tv_dates = sorted(tv_df["Date"].unique().tolist(),
                          key=lambda d: oi_dates.index(d) if d in oi_dates else 999)
        tv_block["dates"] = tv_dates
        for date in tv_dates:
            tv_block["data"][date] = _build_participant_date_block(tv_df, date)

    # ── Day-on-Day block ──────────────────────────────────────────────────
    dod_block = {"pairs": [], "data": {}}
    if len(oi_dates) >= 2:
        for i in range(len(oi_dates) - 1):
            d0, d1 = oi_dates[i], oi_dates[i + 1]
            pair_key = f"{d0}|{d1}"
            dod_block["pairs"].append([d0, d1])
            dod_block["data"][pair_key] = {}
            for part in PARTICIPANTS:
                prev = oi_block["data"][d0].get(part, {})
                curr = oi_block["data"][d1].get(part, {})
                dod_block["data"][pair_key][part] = _dod_record(prev, curr)
            # TV DoD if available
            if (tv_df is not None and
                    d0 in tv_block["data"] and d1 in tv_block["data"]):
                dod_block["data"][pair_key]["_tv"] = {}
                for part in PARTICIPANTS:
                    prev = tv_block["data"][d0].get(part, {})
                    curr = tv_block["data"][d1].get(part, {})
                    dod_block["data"][pair_key]["_tv"][part] = _dod_record(prev, curr)

    # ── Efficiency block ──────────────────────────────────────────────────
    eff_block = {"dates": [], "data": {}}
    if tv_df is not None:
        common_dates = [d for d in oi_dates if d in tv_block["data"]]
        eff_block["dates"] = common_dates
        for date in common_dates:
            eff_block["data"][date] = {}
            for part in PARTICIPANTS:
                oi_rec = oi_block["data"][date].get(part, {})
                tv_rec = tv_block["data"][date].get(part, {})
                eff_block["data"][date][part] = _efficiency_record(oi_rec, tv_rec)

    # ── Sentiment block ───────────────────────────────────────────────────
    sent_block = {"dates": oi_dates, "data": {}}
    for date in oi_dates:
        sent_block["data"][date] = {}
        for part in PARTICIPANTS:
            rec = oi_block["data"][date].get(part, {})
            sent_block["data"][date][part] = {
                "net_fut_idx_dir": "Net Long" if (rec.get("net_fut_idx", 0) or 0) > 0 else "Net Short",
                "net_fut_stk_dir": "Net Long" if (rec.get("net_fut_stk", 0) or 0) > 0 else "Net Short",
                "idx_oi_pcr":      rec.get("idx_oi_pcr"),
                "ls_ratio":        rec.get("ls_ratio"),
                "signal":          _sentiment_signal(rec),
            }

    # ── Chain block ───────────────────────────────────────────────────────
    chain_block = None
    if chain_df is not None and chain_analytics:
        chain_date = list(chain_analytics.keys())[0]
        ana = chain_analytics[chain_date]
        pain_map = ana.get("pain_map", {})

        rows = []
        for _, r in chain_df[chain_df["Date"] == chain_date].iterrows():
            strike = float(r["Strike"])
            c_oi   = float(r.get("C_OI", 0) or 0)
            p_oi   = float(r.get("P_OI", 0) or 0)
            c_iv   = _safe(r.get("C_IV"))
            p_iv   = _safe(r.get("P_IV"))
            rows.append({
                "strike":     strike,
                "c_oi":       int(c_oi),
                "c_chng_oi":  int(float(r.get("C_CHNG_OI", 0) or 0)),
                "c_volume":   int(float(r.get("C_VOLUME",  0) or 0)),
                "c_iv":       round(c_iv, 2) if c_iv is not None else None,
                "c_ltp":      _safe(r.get("C_LTP")),
                "c_chng":     _safe(r.get("C_CHNG")),
                "p_oi":       int(p_oi),
                "p_chng_oi":  int(float(r.get("P_CHNG_OI", 0) or 0)),
                "p_volume":   int(float(r.get("P_VOLUME",  0) or 0)),
                "p_iv":       round(float(p_iv), 2) if p_iv is not None else None,
                "p_ltp":      _safe(r.get("P_LTP")),
                "p_chng":     _safe(r.get("P_CHNG")),
                "pcr":        round(p_oi / c_oi, 4) if c_oi else None,
                "iv_skew":    (round(float(p_iv) - c_iv, 2)
                               if (p_iv is not None and c_iv is not None) else None),
                "combined_oi": int(c_oi + p_oi),
                "pain":       int(pain_map.get(strike, 0)),
            })

        cw = ana.get("call_walls", pd.DataFrame())
        pw = ana.get("put_walls",  pd.DataFrame())

        chain_block = {
            "date":            chain_date,
            "atm":             float(ana["atm"]),
            "max_pain":        float(ana["max_pain"]),
            "oi_pcr":          _safe(ana.get("oi_pcr")),
            "vol_pcr":         _safe(ana.get("vol_pcr")),
            "total_call_oi":   int(ana.get("total_call_oi", 0)),
            "total_put_oi":    int(ana.get("total_put_oi",  0)),
            "total_call_vol":  int(ana.get("total_call_vol", 0)),
            "total_put_vol":   int(ana.get("total_put_vol",  0)),
            "call_walls": (cw[["Strike","C_OI","C_CHNG_OI","C_VOLUME"]]
                             .rename(columns={"Strike":"strike","C_OI":"oi",
                                              "C_CHNG_OI":"chng_oi","C_VOLUME":"volume"})
                             .astype({"strike": float, "oi": int,
                                      "chng_oi": int, "volume": int})
                             .to_dict(orient="records")
                           if not cw.empty else []),
            "put_walls":  (pw[["Strike","P_OI","P_CHNG_OI","P_VOLUME"]]
                             .rename(columns={"Strike":"strike","P_OI":"oi",
                                              "P_CHNG_OI":"chng_oi","P_VOLUME":"volume"})
                             .astype({"strike": float, "oi": int,
                                      "chng_oi": int, "volume": int})
                             .to_dict(orient="records")
                           if not pw.empty else []),
            "data": rows,
        }

    # ── Assemble and write ────────────────────────────────────────────────
    payload = {
        "meta": {
            "generated_at":   datetime.now(timezone.utc).isoformat(),
            "date_tag":       date_tag,
            "latest_date":    oi_dates[-1] if oi_dates else None,
            "oi_dates":       oi_dates,
            "tv_dates":       tv_block["dates"],
            "chain_date":     chain_block["date"] if chain_block else None,
            "excel_filename": excel_filename,
        },
        "participants": PARTICIPANTS,
        "oi":           oi_block,
        "tv":           tv_block,
        "dod":          dod_block,
        "efficiency":   eff_block,
        "sentiment":    sent_block,
        "chain":        chain_block,
    }

    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(payload, f, separators=(",", ":"), allow_nan=False)

    size_kb = Path(out_path).stat().st_size / 1024
    print(f"  data.json written: {out_path}  ({size_kb:.1f} KB)")


def _discover_files(folder: str, prefix: str, n: int) -> list[str]:
    """Return up to n most-recent CSVs matching PREFIXYYYYMMDD*.csv in folder."""
    folder_path = Path(folder)
    pattern = re.compile(rf"^{re.escape(prefix)}(\d{{8}})", re.IGNORECASE)
    matches = []
    for f in folder_path.glob("*.csv"):
        m = pattern.match(f.stem)
        if m:
            matches.append((m.group(1), str(f)))   # (YYYYMMDD, path)
    matches.sort(key=lambda x: x[0])               # oldest → newest
    return [path for _, path in matches[-n:]]


def main():
    parser = argparse.ArgumentParser(
        description="NSE FAO Comprehensive Options Dashboard Generator",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__)
    parser.add_argument(
        "--folder", required=True,
        help="Folder containing FAOOIYYYYMMDD, FAOTVYYYYMMDD, FAOOCYYYYMMDD CSV files")
    parser.add_argument(
        "--json-out", default=None,
        help="Path to write data.json (default: <folder>/../docs/data.json)")

    args = parser.parse_args()
    folder = args.folder
    if not Path(folder).is_dir():
        sys.exit(f"Folder not found: {folder}")

    # ── Discover files ────────────────────────────────────────────────
    oi_files    = _discover_files(folder, "FAOOI", 5)
    tv_files    = _discover_files(folder, "FAOTV", 5)
    chain_files = _discover_files(folder, "FAOOC", 1)

    if not oi_files:
        sys.exit("No FAOOIYYYYMMDD*.csv files found in the folder.")

    print(f"Found {len(oi_files)} OI file(s):    {[Path(f).name for f in oi_files]}")
    if tv_files:
        print(f"Found {len(tv_files)} TV file(s):    {[Path(f).name for f in tv_files]}")
    else:
        print("No TV files found — skipping TV sheets.")
    if chain_files:
        print(f"Found chain file:           {Path(chain_files[0]).name}")
    else:
        print("No OC file found — skipping chain sheets.")

    # ── Output filename: FAOCLAUDEYYYYMMDD.xlsx ───────────────────────
    # Use the YYYYMMDD from the latest OI file
    latest_oi_stem = Path(oi_files[-1]).stem          # e.g. FAOOI20260701
    date_tag_match = re.search(r"(\d{8})", latest_oi_stem)
    date_tag = date_tag_match.group(1) if date_tag_match else "00000000"
    out_path = str(Path(folder) / f"FAOCLAUDE{date_tag}.xlsx")

    # ── Parse OI ─────────────────────────────────────────────────────
    print(f"\nParsing {len(oi_files)} OI file(s)...")
    oi_dates, oi_df = load_participant_files(oi_files)

    # ── Parse TV (align dates to OI dates by position) ───────────────
    tv_row_map = None
    tv_df      = None
    if tv_files:
        # Match TV to OI by date suffix; fall back to positional alignment
        tv_aligned = []
        for oi_f in oi_files:
            oi_date = re.search(r"(\d{8})", Path(oi_f).stem).group(1)
            match = next((f for f in tv_files
                          if re.search(r"(\d{8})", Path(f).stem).group(1) == oi_date), None)
            tv_aligned.append(match)

        # If matched TV files exist for all OI dates, use them; else use latest TV for all
        if all(tv_aligned):
            print(f"Parsing {len(tv_aligned)} TV file(s) (matched by date)...")
            _, tv_df = load_participant_files(tv_aligned, date_overrides=oi_dates)
        else:
            # Use however many TV files exist, aligned positionally to most-recent OI dates
            n_tv = len(tv_files)
            oi_dates_for_tv = oi_dates[-n_tv:]
            print(f"Parsing {n_tv} TV file(s) (positional alignment to last {n_tv} OI dates)...")
            _, tv_df = load_participant_files(tv_files, date_overrides=oi_dates_for_tv)
            # Rebuild oi_dates_for_tv to only include dates that have TV
            # For sheets that need TV, we'll use the matched subset
            # Keep tv_df tagged with the correct oi dates

    # ── Parse Option Chain (single file, tagged with latest OI date) ──
    chain_analytics = {}
    chain_df_all    = None
    chain_dates     = []
    if chain_files:
        print(f"Parsing chain file: {Path(chain_files[0]).name}")
        chain_date, chain_df_all = parse_chain_csv(chain_files[0],
                                                    date_override=oi_dates[-1])
        chain_dates = [chain_date]
        chain_df_all["Date"] = chain_date

        print("Computing chain analytics (max pain, PCR, walls, IV)...")
        sub = chain_df_all.copy()
        mp_strike, pain_map = compute_max_pain(sub)
        pcr_data            = compute_pcr(sub)
        walls               = compute_walls(sub, top_n=10)
        atm                 = compute_atm(sub)
        chain_analytics[chain_date] = {
            "max_pain":  mp_strike,
            "pain_map":  pain_map,
            "atm":       atm,
            **pcr_data,
            **walls,
        }
        print(f"  {chain_date}: Max Pain={mp_strike:,.0f}  "
              f"ATM={atm:,.0f}  OI PCR={pcr_data['oi_pcr']:.3f}")

    # ── Build workbook ────────────────────────────────────────────────
    print("\nBuilding workbook...")
    wb = Workbook()
    wb.remove(wb.active)

    # Raw data sheets
    oi_row_map = write_raw_participant_sheet(wb, "Raw OI Data", oi_dates, oi_df)
    if tv_df is not None:
        tv_row_map = write_raw_participant_sheet(wb, "Raw TV Data", oi_dates, tv_df)
    if chain_df_all is not None:
        write_raw_chain_sheet(wb, chain_dates, chain_df_all)

    # Participant dashboards
    build_oi_dashboard(wb, oi_dates, oi_row_map)
    if tv_row_map:
        build_tv_dashboard(wb, oi_dates, tv_row_map)
    build_net_oi_sheet(wb, oi_dates, oi_row_map)
    build_pcr_sheet(wb, oi_dates, oi_row_map, tv_row_map, chain_analytics)
    if tv_row_map:
        build_efficiency_sheet(wb, oi_dates, oi_row_map, tv_row_map)
    build_dod_sheet(wb, oi_dates, oi_row_map, tv_row_map)

    # Chain dashboards (single-date — pass as single-element list)
    if chain_df_all is not None:
        build_chain_view_sheet(wb, chain_dates, chain_df_all)
        build_max_pain_sheet(wb, chain_dates, chain_df_all, chain_analytics)
        build_oi_walls_sheet(wb, chain_dates, chain_analytics, top_n=10)
        build_iv_skew_sheet(wb, chain_dates, chain_df_all, chain_analytics)

    # Integrated sentiment
    build_sentiment_sheet(wb, oi_dates, oi_row_map, tv_row_map, chain_analytics)


    # ── Export data.json ─────────────────────────────────────────────
    print('\nExporting data.json...')
    if args.json_out:
        json_path = args.json_out
    else:
        json_path = str(Path(folder) / 'docs' / 'data.json')
    Path(json_path).parent.mkdir(parents=True, exist_ok=True)
    export_json(
        out_path        = json_path,
        date_tag        = date_tag,
        excel_filename  = Path(out_path).name,
        oi_dates        = oi_dates,
        oi_df           = oi_df,
        tv_df           = tv_df,
        chain_df        = chain_df_all,
        chain_analytics = chain_analytics,
    )

    wb.save(out_path)
    sheets = [ws.title for ws in wb.worksheets]
    print(f'\n\u2713 Dashboard saved: {out_path}')
    print(f'  {len(sheets)} sheets: {", ".join(sheets)}')

if __name__ == "__main__":
    main()
